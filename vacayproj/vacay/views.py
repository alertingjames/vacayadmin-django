import datetime
import string
from itertools import islice

import xlrd
import re
from django.core import mail
from django.core.mail import send_mail, BadHeaderError, EmailMessage
from django.core.mail import EmailMultiAlternatives
from django.contrib import messages
from _mysql_exceptions import DataError, IntegrityError
from django.template import RequestContext

import time
import io
import requests

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from django.core.files.storage import FileSystemStorage
import json
from django.contrib import auth
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.utils.datastructures import MultiValueDictKeyError
from openpyxl.styles import PatternFill

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.fields import empty
from rest_framework.permissions import AllowAny
from xlrd import XLRDError
from time import gmtime, strftime

from vacay.serializers import AdminUserSerializer, ProviderSerializer, ServiceSerializer, ProviderScheduleSerializer, \
    ProductSerializer, BroadmoorProductSerializer, BroadmoorProductDetailSerializer, EmployeeSerializer, \
    CommonUserSerializer, MailBoxSerializer, WatercoolerSerializer, CommentSerializer
from vacayproj import settings
from .models import AdminUser, Provider, Product, Service, BroadmoorProduct, BroadmoorProductDetail, Employee, Job, \
    Announce, ProviderSchedule, AnnounceView, Account, CommonUser, MailBox, Watercooler, Comment, TipsTricks, Img
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.sessions.backends.db import SessionStore
from django.contrib.sessions.models import Session
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate
from django import forms
import sys

pro_num=0
serv=0
serv2=0
prod=0
ret=0
com=0
emp=0
jb=0
ann=0

import pyrebase

config = {
    "apiKey": "AIzaSyC1e8-93-ia4seBJlMR7_BPt-pxnWGrgNA",
    "authDomain": "vacay-42bcd.firebaseapp.com",
    "databaseURL": "https://vacay-42bcd.firebaseio.com",
    "storageBucket": "vacay-42bcd.appspot.com"
}

firebase = pyrebase.initialize_app(config)

class UploadFileForm(forms.Form):
    file = forms.FileField()

def index(request):
    return HttpResponse('<h2>Hello VaCay!</h2>')

def login_user_view(request):
    global serv
    global ret
    global com
    global serv2
    global prod
    global emp
    global jb
    global ann
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return render(request, 'vacay/login_user.html', {})
    else:
        adminUser = AdminUser.objects.get(adminID=idx)
        adminType = adminUser.adminBroadmoor

        adminUsers = AdminUser.objects.filter(adminBroadmoor='1')
        adminCount = adminUsers.count()
        if adminCount > 10:
            adminUsers = islice(reversed(adminUsers), 0, 9)
        else:
            adminUsers = islice(reversed(adminUsers), 0, adminCount)

        providers = Provider.objects.all()
        services = Service.objects.all()
        products = Product.objects.all()
        broadmoorProducts = BroadmoorProduct.objects.all()
        employees = Employee.objects.all()
        jobs = Job.objects.all()
        announces = Announce.objects.all()

        providerCount = providers.count()
        serviceCount = services.count()
        productCount = products.count()
        brproductCount = broadmoorProducts.count()
        employeeCount = employees.count()
        jobCount = jobs.count()
        announceCount = announces.count()

        if providerCount > 10:
            providers = islice(reversed(providers), 0, 9)
        else:
            providers = islice(reversed(providers), 0, providerCount)

        if serviceCount > 10:
            services = islice(reversed(services), 0, 9)
        else:
            services = islice(reversed(services), 0, serviceCount)

        if productCount > 10:
            products = islice(reversed(products), 0, 9)
        else:
            products = islice(reversed(products), 0, productCount)

        if brproductCount > 10:
            broadmoorProducts = islice(reversed(broadmoorProducts), 0, 9)
        else:
            broadmoorProducts = islice(reversed(broadmoorProducts), 0, brproductCount)

        if employeeCount > 10:
            employees = islice(reversed(employees), 0, 9)
        else:
            employees = islice(reversed(employees), 0, employeeCount)

        if jobCount > 10:
            jobs = islice(reversed(jobs), 0, 9)
        else:
            jobs = islice(reversed(jobs), 0, jobCount)

        if announceCount > 10:
            announces = islice(reversed(announces), 0, 9)
        else:
            announces = islice(reversed(announces), 0, announceCount)

        # return HttpResponse(adminUser.adminBroadmoor)

        # flags = Flags.objects.filter(adminID=idx)         #/////////////////////////////////////
        # count=flags.count()
        # if count == 0:
        #     flag=Flags()
        #     flag.adminID=idx
        #     flag.save()
        # flags = Flags.objects.filter(adminID=idx)

        if adminType == "0":

            user_id = adminUser.adminID
            serv=1
            serv2=0
            prod=0

            # flags[0].fservice = 1                       #/////////////////////////////////////////
            # flags[0].fservice2 = 0
            # flags[0].fproduct = 0
            # flags[0].save()

            request.session['serv']=1
            request.session['serv2']=0
            request.session['prod']=0

            context = {'providers': providers, 'services': services, 'products': products}
            return render(request, 'vacay/home.html', context)

        elif adminType == "1":
            ret=1
            request.session['ret']=1
            user_id = adminUser.adminID

            # flags[0].fretail = 1                          #/////////////////////////////////////////
            # flags[0].save()

            context = {'broadmoors': broadmoorProducts, 'adminUsers': adminUsers, 'products': products}
            return render(request, 'vacay/home_broadmoor.html', context)
            # return HttpResponse(adminType)

        elif adminType == "2":
            com=1
            request.session['com']=1
            user_id = adminUser.adminID

            # flags[0].fcompany = 1                        #/////////////////////////////////////
            # flags[0].save()

            context = {'employees': employees, 'jobs': jobs, 'announces': announces}
            return render(request, 'vacay/home_company.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def upload_adminpicture(request):
    if request.method == 'POST':
        image = request.POST.get('b64', None)
        adminid = request.POST.get('adminID', None)
        adminUser = AdminUser.objects.get(id=adminid)
        adminUser.adminImageUrl = image
        adminUser.save()
        resp = {'result_code':'0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def register_admin(request):

    if request.method == 'POST':

        eml = request.POST.get('adminEmail', '')
        name = request.POST.get('adminName', '')
        password = request.POST.get('adminPassword', '')
        broadmoor = request.POST.get('adminBroadmoor', '')
        company = request.POST.get('adminCompany', '')

        users = AdminUser.objects.filter(adminEmail=eml)
        count = users.count()

        if count ==0:

            adminUser = AdminUser()
            adminUser.adminEmail = eml
            adminUser.adminName = name
            adminUser.adminPassword = password
            adminUser.adminBroadmoor = broadmoor
            if company:
                adminUser.adminCompany = company
            else:
                adminUser.adminCompany = ""

            adminUser.save()

            adminUser.adminID = adminUser.pk
            adminUser.save()

            user1 = User()
            user1.username = eml
            user1.email = eml
            user1.password = password
            user1.set_password(password)
            user1.save()

            user = authenticate(username=eml, password=password)
            login(request,user)

            resp = {'result_code': '0', 'adminID': adminUser.pk}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:

            resp_er = {'result_code': '101'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def upload__admin_image(request):

    if request.method == 'POST':

        image = request.FILES['file']

        fs = FileSystemStorage()
        filename = fs.save(image.name, image)
        uploaded_file_url = fs.url(filename)

        adminId = request.POST.get('adminID')
        adminUser = AdminUser.objects.get(id=adminId)
        adminUser.adminImageUrl = settings.URL + uploaded_file_url
        adminUser.save()

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)


    elif request.method == 'GET':

        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def upload__broadmoor_image(request):

    if request.method == 'POST':
        image = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(image.name, image)
        uploaded_file_url = fs.url(filename)

        adminId = request.POST.get('adminID')
        adminUser = AdminUser.objects.get(id=adminId)
        adminUser.adminLogoImageUrl = settings.URL + uploaded_file_url
        adminUser.save()

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)


    elif request.method == 'GET':

        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def register_company_admin(request):

    if request.method == 'POST':

        eml = request.POST.get('adminEmail', None)
        name = request.POST.get('adminName', None)
        password = request.POST.get('adminPassword', None)
        broadmoor = request.POST.get('adminBroadmoor', None)
        company = request.POST.get('adminCompany', None)

        users = AdminUser.objects.filter(adminEmail=eml)
        count = users.count()

        if count ==0:

            adminUser = AdminUser()
            adminUser.adminEmail = eml
            adminUser.adminName = name
            adminUser.adminPassword = password
            adminUser.adminBroadmoor = broadmoor
            adminUser.adminCompany = company

            adminUser.save()

            adminUser.adminID = adminUser.pk
            adminUser.save()

            user1 = User()
            user1.username = eml
            user1.email = eml
            user1.password = password
            user1.set_password(password)
            user1.save()

            user = authenticate(username=eml, password=password)
            login(request, user)

            resp = {'result_code': '0', 'adminID': adminUser.pk}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:

            resp_er = {'result_code': '101'}
            return HttpResponse(json.dumps(resp_er))

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_admin_data(request, admin_id):

    if request.method == 'GET':

        adminUser = AdminUser.objects.filter(adminID=admin_id)
        count = adminUser.count()
        if count>0:
            serializer = AdminUserSerializer(adminUser, many=True)
            resp = {'result_code': '0', 'adminData': serializer.data, }
            return JsonResponse(resp, status=status.HTTP_200_OK)
        else:
            resp = {'result_code': '113'}
            return JsonResponse(resp)
    elif request.method == 'POST':
        pass



@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login_admin_from_app(request):

    if request.method == 'POST':

        email = request.POST.get('adminEmail', None)
        password = request.POST.get('adminPassword', None)

        user0 = AdminUser.objects.filter(adminEmail=email)
        count = user0.count()
        if count>0:

            user = authenticate(username=email, password=password)
            if user is not None:
                login(request, user)
            else:
                resp = {'result_code': '113'}
                return JsonResponse(resp, status=status.HTTP_400_BAD_REQUEST)

            serializer = AdminUserSerializer(user0, many=True)
            resp = {'result_code': '0', 'adminData': serializer.data}

            return JsonResponse(resp, status=status.HTTP_200_OK)

            # user = authenticate(username=email, password=password)
            # if user is not None:
            #     login(request, user)
            #     if not request.session.exists(request.session.session_key):
            #         request.session.create()
            #     session_key = request.session.session_key
            #     session = Session.objects.get(session_key=session_key)
            #     idx = session.get_decoded().get('_auth_user_id')
            #     user = User.objects.get(pk=idx)
            #     eml = user.email
            #
            #     serializer = AdminUserSerializer(user0, many=True)
            #     resp = {'result_code': '0', 'adminData': serializer.data}
            #
            #     return JsonResponse(resp, status=status.HTTP_200_OK)
            #
            # else:
            #     resp = {'result_code': '113'}
            #     return JsonResponse(resp, status=status.HTTP_400_BAD_REQUEST)
        else:
            resp = {'result_code': '113'}
            return JsonResponse(resp, status=status.HTTP_400_BAD_REQUEST)


    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login_admin_from_web(request):

    if request.method == 'POST':

        email = request.POST.get('email', None)
        password = request.POST.get('password', None)

        user = authenticate(username=email, password=password)
        if user is not None:
            login(request, user)
            if not request.session.exists(request.session.session_key):
                request.session.create()
            session_key = request.session.session_key
            session = Session.objects.get(session_key=session_key)
            idx = session.get_decoded().get('_auth_user_id')

            # return HttpResponse(idx)

            # return HttpResponse(email)

            adminUser = AdminUser.objects.get(adminEmail=email)
            adminType = adminUser.adminBroadmoor

            adminUsers = AdminUser.objects.filter(adminBroadmoor='1')
            adminCount = adminUsers.count()
            if adminCount > 10:
                adminUsers = islice(reversed(adminUsers), 0, 9)
            else:
                adminUsers = islice(reversed(adminUsers), 0, adminCount)

            providers = Provider.objects.all()
            services = Service.objects.all()
            products = Product.objects.all()
            broadmoorProducts = BroadmoorProduct.objects.all()
            employees = Employee.objects.all()
            jobs = Job.objects.all()
            announces = Announce.objects.all()

            providerCount = providers.count()
            serviceCount = services.count()
            productCount = products.count()
            brproductCount = broadmoorProducts.count()
            employeeCount = employees.count()
            jobCount = jobs.count()
            announceCount = announces.count()

            if providerCount > 10:
                providers = islice(reversed(providers), 0, 9)
            else:
                providers = islice(reversed(providers), 0, providerCount)

            if serviceCount > 10:
                services = islice(reversed(services), 0, 9)
            else:
                services = islice(reversed(services), 0, serviceCount)

            if productCount > 10:
                products = islice(reversed(products), 0, 9)
            else:
                products = islice(reversed(products), 0, productCount)

            if brproductCount > 10:
                broadmoorProducts = islice(reversed(broadmoorProducts), 0, 9)
            else:
                broadmoorProducts = islice(reversed(broadmoorProducts), 0, brproductCount)

            if employeeCount > 10:
                employees = islice(reversed(employees), 0, 9)
            else:
                employees = islice(reversed(employees), 0, employeeCount)

            if jobCount > 10:
                jobs = islice(reversed(jobs), 0, 9)
            else:
                jobs = islice(reversed(jobs), 0, jobCount)

            if announceCount > 10:
                announces = islice(reversed(announces), 0, 9)
            else:
                announces = islice(reversed(announces), 0, announceCount)

            # return HttpResponse(adminUser.adminBroadmoor)

            if adminType == "0":
                user_id= adminUser.adminID

                request.session['serv'] = 1
                request.session['serv2'] = 0
                request.session['prod'] = 0

                context = {'providers':providers, 'services':services, 'products':products}
                return render(request, 'vacay/home.html', context)
            elif adminType == "1":
                user_id = adminUser.adminID

                request.session['ret'] = 1

                context = {'broadmoors':broadmoorProducts, 'adminUsers': adminUsers, 'products':products}
                return render(request, 'vacay/home_broadmoor.html', context)
                # return HttpResponse(adminType)
            elif adminType == "2":
                user_id = adminUser.adminID

                request.session['com'] = 1

                context = {'employees':employees, 'jobs':jobs, 'announces':announces}
                return render(request, 'vacay/home_company.html', context)
            else:
                if len(email) > 0:
                    note = "note"
                else:
                    note = ""
                return render(request, 'vacay/login_user.html', {'note': note})

        else:
            if len(email)>0:
                note="note"
            else:
                note = ""
            return render(request, 'vacay/login_user.html', {'note': note})


    elif request.method == 'GET':

        pass

def logout(request):
    auth.logout(request)
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    # flags = Flags.objects.filter(adminID=idx)

    request.session['serv'] = 1
    request.session['serv2'] = 0
    request.session['prod'] = 0

    if idx is None:
        return render(request, 'vacay/login_user.html', {})
    return render(request, 'vacay/login_user.html', {})

def get_all_providers(request):
    global serv
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    adminUser=AdminUser.objects.get(adminID=idx)
    all_providers=Provider.objects.filter(adminID=idx).order_by('-id')

    serv=0
    request.session['serv']=0

    context = {'all_providers': all_providers, 'admin': adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/index.html', context)

    # return HttpResponse(idx)

def get_services(request, provider_id):

    global pro_num
    global serv2
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    request.session['pro_id'] = provider_id

    provider = Provider.objects.get(proid=provider_id)
    services=Service.objects.filter(proid=provider_id).order_by('-id')

    # services = Service.objects.all()
    serv2=request.session['serv2']

    context = {'services': services, 'provider':provider}
    # cache.set('search_env', 3)
    if serv2==1:
        return render(request, 'vacay/service_list.html', context)
    else:
        return render(request, 'vacay/show_services.html', context)

def get_all_services(request):

    global pro_num
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    pro_num=request.session['pro_id']

    provider = Provider.objects.get(proid=pro_num)
    services = Service.objects.filter(proid=pro_num).order_by('-id')
    context = {'services': services, 'provider': provider}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_services.html', context)
    # return HttpResponse(pro_num)


def get_products(request, provider_id):

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    request.session['pro_id'] = provider_id

    provider = Provider.objects.get(proid=provider_id)
    product = Product.objects.filter(proid=provider_id).order_by('-id')
    context = {'product': product, 'provider': provider}
    # cache.set('search_env', 3)
    prod=request.session['prod']
    if prod==1:
        return render(request, 'vacay/product_list.html', context)
    else:
        return render(request, 'vacay/show_products.html', context)


def get_all_products(request):

    global pro_num
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    pro_num=request.session['pro_id']

    provider = Provider.objects.get(proid=pro_num)
    product = Product.objects.filter(proid=pro_num).order_by('-id')
    context = {'product': product, 'provider': provider}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_products.html', context)

def get_setup(request, provider_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    adminUser = AdminUser.objects.get(adminID=idx)
    provider = Provider.objects.filter(proid=provider_id)
    context = {'provider': provider, 'admin':adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_provider_setup.html', context)

def get_setups(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    adminUser = AdminUser.objects.get(adminID=idx)
    provider = Provider.objects.filter(adminID=idx)
    context = {'provider': provider, 'admin':adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_provider_setup.html', context)

def edit_provider_view(request, provider_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    provider = Provider.objects.get(proid=provider_id)
    context={'provider': provider}
    return render(request, 'vacay/edit_provider.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edit_provider(request, provider_id):
    if request.method == 'POST':
        firstName = request.POST.get('firstname', None)
        lastName = request.POST.get('lastname', None)
        email = request.POST.get('email', None)
        password = request.POST.get('password', None)
        city = request.POST.get('city', None)
        address = request.POST.get('address', None)
        company = request.POST.get('company', None)
        phone = request.POST.get('phone', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        provider = Provider.objects.get(proid=provider_id)
        provider.proFirstName = firstName
        provider.proLastName = lastName
        provider.proEmail = email
        provider.proPassword = password
        provider.proPhone = phone
        provider.proCity = city
        provider.proAddress = address
        provider.proCompany = company

        try:
            image = request.FILES['photo']
            fs = FileSystemStorage()
            try:
                x = request.POST.get('x', '0')
                y = request.POST.get('y', '0')
                w = request.POST.get('w', '32')
                h = request.POST.get('h', '32')
                #  return HttpResponse(w)
                file = profile_process(image, x, y, w, h)
                image = file
                # return HttpResponse('Cropped!')

            except MultiValueDictKeyError:
                print('No cropping')
            except ValueError:
                print('No cropping')

            filename = fs.save(image.name, image)
            uploaded_file_url = fs.url(filename)
            provider.proProfileImageUrl = settings.URL + uploaded_file_url

        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        provider.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        all_providers = Provider.objects.filter(adminID=idx).order_by('-id')
        context = {'all_providers': all_providers, 'admin':adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/index.html', context)

    elif request.method == 'GET':
        pass

def delete_provider(request, provider_id):
    Provider.objects.get(proid=provider_id).delete()
    Service.objects.filter(proid=provider_id).delete()
    Product.objects.filter(proid=provider_id).delete()
    ProviderSchedule.objects.filter(proid=provider_id).delete()

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    adminUser = AdminUser.objects.get(adminID=idx)
    all_providers = Provider.objects.filter(adminID=idx).order_by('-id')
    context = {'all_providers': all_providers, 'admin':adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/index.html', context)

def add_provider_view(request):
    return render(request, 'vacay/add_provider.html', {})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_provider(request):
    if request.method == 'POST':
        firstName = request.POST.get('firstname', None)
        lastName = request.POST.get('lastname', None)
        email = request.POST.get('email', None)
        password = request.POST.get('password', None)
        phone = request.POST.get('phone', None)
        city = request.POST.get('city', None)
        address = request.POST.get('address', None)
        company = request.POST.get('company', None)
        servicepercent = request.POST.get('servicepercent', None)
        salaryamount = request.POST.get('salaryamount', None)
        productpercent = request.POST.get('productpercent', None)

        image = request.FILES['photo']
        fs = FileSystemStorage()
        try:
            x = request.POST.get('x', '0')
            y = request.POST.get('y', '0')
            w = request.POST.get('w', '32')
            h = request.POST.get('h', '32')
            #  return HttpResponse(w)
            file = profile_process(image, x, y, w, h)
            image = file
            # return HttpResponse('Cropped!')

        except MultiValueDictKeyError:
            print('No cropping')
        except ValueError:
            print('No cropping')

        filename = fs.save(image.name, image)
        uploaded_file_url = fs.url(filename)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        provider=Provider()
        provider.adminID = idx
        provider.proFirstName = firstName
        provider.proLastName = lastName
        provider.proEmail = email
        provider.proPassword = password
        provider.proPhone = phone
        provider.proCity = city
        provider.proAddress = address
        provider.proCompany = company
        provider.proServicePercent = servicepercent
        provider.proSalary = salaryamount
        provider.proProductSalePercent = productpercent

        provider.proProfileImageUrl = settings.URL + uploaded_file_url

        provider.save()
        provider.proid = provider.pk
        provider.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        all_providers = Provider.objects.filter(adminID=idx).order_by('-id')
        context = {'all_providers': all_providers, 'admin':adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/index.html', context)

    elif request.method == 'GET':
        pass

from PIL import Image
from django.core.files.uploadedfile import *

def profile_process(image, x, y, w, h):
    #      return HttpResponse(w)
    x = float(x)
    y = float(y)
    w = float(w)
    h = float(h)
    #     return HttpResponse(w)
    file = None
    try:
        thumb_io = io.BytesIO()
        image_file = Image.open(image)
   #     resized_image = image_file.resize((600, int(250 * image_file.height / image_file.width)), Image.ANTIALIAS)
        cropped_image = image_file.crop((x, y, w + x, h + y))
   #     resized_image = cropped_image.resize((160, 160), Image.ANTIALIAS)
        cropped_image.save(thumb_io, image.content_type.split('/')[-1].upper())

        # creating new InMemoryUploadedFile() based on the modified file
        file = InMemoryUploadedFile(thumb_io,
                                    u"photo",  # important to specify field name here
                                    "croppedimage.jpg",
                                    image.content_type,
                                    None,
                                    None)
    except OSError:
        print('Invalid file!')

    return file


def edit_setup_view(request, provider_id):

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        provider = Provider.objects.get(proid=provider_id)
        context = {'provider': provider}
        # cache.set('search_env', 3)
        return render(request, 'vacay/edit_setup.html', context)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edit_setup(request, provider_id):

    if request.method == 'POST':

        servicePercent = request.POST.get('servicepercent', None)
        salary = request.POST.get('salaryamount', None)
        productPercent = request.POST.get('productpercent', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        provider = Provider.objects.get(proid=provider_id)

        provider.proServicePercent = servicePercent
        provider.proSalary = salary
        provider.proProductSalePercent = productPercent

        provider.save()
        provider = Provider.objects.filter(proid=provider_id)
        context = {'provider': provider}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_provider_setup.html', context)
        # return HttpResponse(provider.proid)

    elif request.method == 'GET':
        pass

def edit_service_view(request, service_id, provider_id):

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    service = Service.objects.get(serviceid=service_id)
    provider = Provider.objects.get(proid=provider_id)
    context={'service':service, 'provider':provider}
    # cache.set('search_env', 3)
    return render(request, 'vacay/edit_service.html', context)

def delete_service(request, service_id, provider_id):

    global pro_num
    Service.objects.get(serviceid=service_id).delete()

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    provider = Provider.objects.get(proid=provider_id)
    services = Service.objects.filter(proid=provider_id).order_by('-id')
    context = {'services': services, 'provider': provider}

    return render(request, 'vacay/show_services.html', context)

def add_service_view(request):
    global serv2
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    serv2=request.session['serv2']

    if serv2==1:
        return render(request, 'vacay/add_home_service.html')
    else:
        return render(request, 'vacay/add_service.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_service(request):

    global pro_num
    global serv2

    if request.method == 'POST':

        category = request.POST.get('category', None)
        serviceName = request.POST.get('servicename', None)
        servicePrice = request.POST.get('serviceprice', None)
        description = request.POST.get('servicedescription', None)
        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        pro_num=request.session['pro_id']
        serv2=request.session['serv2']

        service = Service()
        service.adminID = idx
        service.proid = pro_num
        service.proBeautyCategory = category
        service.proBeautySubCategory = serviceName
        if "$" in servicePrice:
            service.proServicePrice = servicePrice
        else:
            service.proServicePrice = "$"+servicePrice
        service.proServiceDescription = description

        try:
            image = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            service.proServicePictureUrl = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            service.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            service.youtube_url=youtubeurl

        service.save()
        service.serviceid = service.pk
        service.save()

        provider = Provider.objects.get(proid=pro_num)
        services = Service.objects.filter(proid=pro_num).order_by('-id')
        context = {'services': services, 'provider':provider}
        # cache.set('search_env', 3)

        if serv2==1:
            return render(request, 'vacay/service_list.html', context)
        else:
            return render(request, 'vacay/show_services.html', context)


    elif request.method == 'GET':
        pass


def edit_product_view(request, product_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    product = Product.objects.get(itemid=product_id)
    context = {'product': product}
    return render(request, 'vacay/edit_product.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edit_service(request, service_id, provider_id ):

    global pro_num
    if request.method == 'POST':
        category = request.POST.get('category', None)
        serviceName = request.POST.get('servicename', None)
        servicePrice = request.POST.get('serviceprice', None)
        serviceDescription = request.POST.get('servicedescription', None)
        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        service = Service.objects.get(serviceid=service_id)
        service.proBeautyCategory = category
        service.proBeautySubCategory = serviceName
        if "$" in servicePrice:
            service.proServicePrice = servicePrice
        else:
            service.proServicePrice = "$"+str(servicePrice)
        service.proServiceDescription = serviceDescription

        try:
            photo = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            service.proServicePictureUrl = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            service.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        service.youtube_url = youtubeurl

        service.save()
        provider = Provider.objects.get(proid=provider_id)
        services = Service.objects.filter(proid=provider_id).order_by('-id')
        context = {'services': services, 'provider':provider}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_services.html', context)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def editservicepictures(request, service_id, provider_id ):

    global pro_num
    if request.method == 'POST':

        service = Service.objects.get(serviceid=service_id)

        try:
            image = request.FILES['photoa']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            service.imageA = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photob']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            service.imageB = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photoc']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            service.imageC = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photod']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            service.imageD = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photoe']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            service.imageE = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photof']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            service.imageF = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        desca = request.POST.get('desca', '')
        descb = request.POST.get('descb', '')
        descc = request.POST.get('descc', '')
        descd = request.POST.get('descd', '')
        desce = request.POST.get('desce', '')
        descf = request.POST.get('descf', '')

        service.descA=desca
        service.descB=descb
        service.descC=descc
        service.descD=descd
        service.descE=desce
        service.descF=descf

        service.save()
        provider = Provider.objects.get(proid=provider_id)
        services = Service.objects.filter(proid=provider_id).order_by('-id')
        context = {'services': services, 'provider':provider}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_services.html', context)


def add_product_view(request):
    global prod
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    prod=request.session['prod']
    if prod==1:
        return render(request, 'vacay/add_home_product.html')
    else:
        return render(request, 'vacay/add_product.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edit_product(request, product_id):

    global pro_num
    if request.method == 'POST':
        product_t = request.POST.get('product', None)
        productName = request.POST.get('productname', None)
        brand = request.POST.get('brand', None)
        size = request.POST.get('size', None)
        price = request.POST.get('price', None)
        inventory = request.POST.get('inventory', None)
        saleStatus = request.POST.get('salestatus', None)
        productDescription = request.POST.get('productdescription', None)
        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        product = Product.objects.get(itemid=product_id)
        product.itemProduct = product_t
        product.itemName = productName
        product.itemBrand = brand
        product.itemSize = size
        if "$" in price:
            product.itemPrice = price
        else:
            product.itemPrice = "$" + str(price)
        product.itemInventoryNum = inventory
        product.itemDescription = productDescription
        product.itemSaleStatus = saleStatus

        try:
            photo = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            product.itemPictureUrl = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            product.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            product.youtube_url = youtubeurl

        pro_num=request.session['pro_id']

        product.save()
        provider = Provider.objects.get(proid=pro_num)
        product = Product.objects.filter(proid=pro_num).order_by('-id')
        context = {'product': product, 'provider':provider}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_products.html', context)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def editproductpictures(request, product_id):

    global pro_num
    if request.method == 'POST':
        product = Product.objects.get(itemid=product_id)
        try:
            photo = request.FILES['photoa']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            product.imageA = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photob']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            product.imageB = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photoc']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            product.imageC = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photod']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            product.imageD = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photoe']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            product.imageE = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photof']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            product.imageF = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        product.descA = request.POST.get('desca', '')
        product.descB = request.POST.get('descb', '')
        product.descC = request.POST.get('descc', '')
        product.descD = request.POST.get('descd', '')
        product.descE = request.POST.get('desce', '')
        product.descF = request.POST.get('descf', '')

        pro_num=request.session['pro_id']

        product.save()
        provider = Provider.objects.get(proid=pro_num)
        product = Product.objects.filter(proid=pro_num).order_by('-id')
        context = {'product': product, 'provider':provider}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_products.html', context)

    elif request.method == 'GET':
        pass

def delete_product(request, product_id):
    global pro_num
    Product.objects.get(itemid=product_id).delete()

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    pro_num = request.session['pro_id']

    provider = Provider.objects.get(proid=pro_num)
    product = Product.objects.filter(proid=pro_num).order_by('-id')
    context = {'product': product, 'provider': provider}

    return render(request, 'vacay/show_products.html', context)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_product(request):

    global pro_num
    global prod
    if request.method == 'POST' and request.FILES['photo']:

        product_t = request.POST.get('product', None)
        productName = request.POST.get('productname', None)
        brand = request.POST.get('brand', None)
        size = request.POST.get('size', None)
        price = request.POST.get('price', None)
        inventory = request.POST.get('inventory', None)
        saleStatus = request.POST.get('salestatus', None)
        productDescription = request.POST.get('productdescription', None)
        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        pro_num = request.session['pro_id']

        product = Product()
        product.proid = pro_num
        product.itemProduct = product_t
        product.itemName = productName
        product.itemBrand = brand
        product.itemSize = size
        if "$" in price:
            product.itemPrice = price
        else:
            product.itemPrice = "$" + str(price)
        product.itemInventoryNum = inventory
        product.itemDescription = productDescription
        product.itemSaleStatus = saleStatus
        try:
            photo = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            product.itemPictureUrl = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            product.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            product.youtube_url = youtubeurl

        product.save()
        product.itemid = product.pk
        product.save()

        provider = Provider.objects.get(proid=pro_num)
        product = Product.objects.filter(proid=pro_num).order_by('-id')
        context = {'product': product, 'provider':provider}

        prod = request.session['prod']

        # cache.set('search_env', 3)
        if prod==1:
            return render(request, 'vacay/product_list.html', context)
        else:
            return render(request, 'vacay/show_products.html', context)

    elif request.method == 'GET':
        pass


def get_broadmoor_products(request):
    global ret
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    ret=0
    request.session['ret']=0
    adminUser = AdminUser.objects.get(adminID=idx)
    bproducts=BroadmoorProduct.objects.filter(adminID=idx).order_by('-id')
    context = {'bproducts': bproducts, 'admin':adminUser }
    return render(request, 'vacay/show_broadmoor_products.html', context)
    # return HttpResponse(idx)

def add_broadmoor_product(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    return render(request, 'vacay/add_broadmoor_product.html')

def add_broadmoor(request):
    if request.method == 'POST' and request.FILES['photo']:

        productName = request.POST.get('productname', None)
        inventory = request.POST.get('inventory', None)
        category = request.POST.get('category', None)
        additional = request.POST.get('additionalmaterial', None)
        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        bproduct = BroadmoorProduct()
        bproduct.adminID = idx
        bproduct.bm_proName = productName
        bproduct.bm_proInventoryNum = inventory
        bproduct.bm_proCategory = category
        bproduct.bm_proAdditional = additional

        try:
            photo = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            bproduct.bm_proImageUrl = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            bproduct.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            bproduct.youtube_url = youtubeurl

        bproduct.save()
        bproduct.bm_proid = bproduct.pk
        bproduct.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        bproducts = BroadmoorProduct.objects.filter(adminID=idx).order_by('-id')
        context = {'bproducts': bproducts, 'admin':adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_broadmoor_products.html', context)

    elif request.method == 'GET':
        pass


def edit_broadmoor_product(request, broadmoorproduct_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    bproduct=BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id)
    context = {'bproduct': bproduct}
    # cache.set('search_env', 3)
    return render(request, 'vacay/edit_broadmoor_product.html', context)

def update_broadmoor_product(request, broadmoorproduct_id):

    if request.method == 'POST':

        productName = request.POST.get('productname', None)
        inventory = request.POST.get('inventory', None)
        category = request.POST.get('category', None)
        additional = request.POST.get('additionalmaterial', None)
        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        bproduct = BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id)

        bproduct.bm_proName = productName
        bproduct.bm_proInventoryNum = inventory
        bproduct.bm_proCategory = category
        bproduct.bm_proAdditional = additional

        try:
            photo = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            bproduct.bm_proImageUrl = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            bproduct.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            bproduct.youtube_url = youtubeurl

        bproduct.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        bproducts = BroadmoorProduct.objects.filter(adminID=idx).order_by('-id')
        context = {'bproducts': bproducts, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_broadmoor_products.html', context)

    elif request.method == 'GET':
        pass

def updatebroadmoorpictures(request, broadmoorproduct_id):

    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        bproduct = BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id)

        try:
            photo = request.FILES['photoa']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            bproduct.imageA = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photob']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            bproduct.imageB = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photoc']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            bproduct.imageC = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photod']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            bproduct.imageD = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photoe']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            bproduct.imageE = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            photo = request.FILES['photof']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            bproduct.imageF = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        bproduct.descA = request.POST.get('desca', '')
        bproduct.descB = request.POST.get('descb', '')
        bproduct.descC = request.POST.get('descc', '')
        bproduct.descD = request.POST.get('descd', '')
        bproduct.descE = request.POST.get('desce', '')
        bproduct.descF = request.POST.get('descf', '')

        bproduct.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        bproducts = BroadmoorProduct.objects.filter(adminID=idx).order_by('-id')
        context = {'bproducts': bproducts, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_broadmoor_products.html', context)

    elif request.method == 'GET':
        pass


def detail_broadmoor_product(request, broadmoorproduct_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    adminUser = AdminUser.objects.get(adminID=idx)
    bproduct=BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id)
    bdetails = BroadmoorProductDetail.objects.filter(bm_proid=broadmoorproduct_id)
    context = {'bdetails': bdetails, 'bproduct':bproduct, 'admin':adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/detail_broadmoor_product.html', context)

def add_detail_broadmoor(request, broadmoorproduct_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    adminUser = AdminUser.objects.get(adminID=idx)
    bproduct=BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id)
    context = {'bproduct': bproduct, 'admin':adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/add_broadmoor_detail.html', context)

def add_broadmoor_detail(request, broadmoorproduct_id):

    if request.method == 'POST':

        size = request.POST.get('size', None)
        quantity = request.POST.get('quantity', None)
        price = request.POST.get('price', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        bdetail = BroadmoorProductDetail()
        bdetail.bm_proid = broadmoorproduct_id
        bdetail.bm_proSize = size
        bdetail.bm_proQuantity = quantity

        if "$" in price:
            bdetail.bm_proPrice = price
        else:
            bdetail.bm_proPrice = "$"+str(price)

        bdetail.save()
        bdetail.bm_detailID = bdetail.pk
        bdetail.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        bproduct = BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id)
        bdetails = BroadmoorProductDetail.objects.filter(bm_proid=broadmoorproduct_id)
        context = {'bdetails': bdetails, 'bproduct': bproduct, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/detail_broadmoor_product.html', context)

    elif request.method == 'GET':
        pass

def edit_detail_broadmoor(request, broadmoorproductdetail_id, broadmoorproduct_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    adminUser = AdminUser.objects.get(adminID=idx)
    bproduct = BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id)
    bdetail = BroadmoorProductDetail.objects.get(bm_detailID=broadmoorproductdetail_id)
    context = {'bdetail': bdetail, 'admin':adminUser, 'bproduct':bproduct}
    # cache.set('search_env', 3)
    return render(request, 'vacay/edit_broadmoor_detail.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edit_detail_broadmoor_product(request, broadmoorproduct_id, broadmoorproductdetail_id):

    if request.method == 'POST':

        size = request.POST.get('size', None)
        quantity = request.POST.get('quantity', None)
        price = request.POST.get('price', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        bdetail = BroadmoorProductDetail.objects.get(bm_detailID=broadmoorproductdetail_id)

        bdetail.bm_proSize = size
        bdetail.bm_proQuantity = quantity

        if "$" in price:
            bdetail.bm_proPrice = price
        else:
            bdetail.bm_proPrice = "$"+str(price)

        bdetail.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        bproduct = BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id)
        bdetails = BroadmoorProductDetail.objects.filter(bm_proid=broadmoorproduct_id)
        context = {'bdetails': bdetails, 'bproduct': bproduct, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/detail_broadmoor_product.html', context)


    elif request.method == 'GET':
        pass

def delete_detail_broadmoor(request, broadmoorproductdetail_id, broadmoorproduct_id):
    BroadmoorProductDetail.objects.get(bm_detailID=broadmoorproductdetail_id).delete()

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    adminUser = AdminUser.objects.get(adminID=idx)
    bproduct = BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id)
    bdetails = BroadmoorProductDetail.objects.filter(bm_proid=broadmoorproduct_id)
    context = {'bdetails': bdetails, 'bproduct': bproduct, 'admin': adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/detail_broadmoor_product.html', context)

def delete_broadmoor_product(request, broadmoorproduct_id):

    BroadmoorProduct.objects.get(bm_proid=broadmoorproduct_id).delete()
    BroadmoorProductDetail.objects.filter(bm_proid=broadmoorproduct_id).delete()

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    bproducts = BroadmoorProduct.objects.filter(adminID=idx).order_by('-id')
    context = {'bproducts': bproducts, 'admin': adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_broadmoor_products.html', context)

def get_employees(request):
    global com
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    com=0
    request.session['com']=0

    adminUser = AdminUser.objects.get(adminID=idx)

    employees=Employee.objects.filter(adminID=idx).order_by('-id')
    context = {'employees': employees, 'admin':adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_employees.html', context)

def edit_employee(request, employee_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    employee=Employee.objects.get(em_id=employee_id)
    context = {'employee': employee}
    # cache.set('search_env', 3)
    return render(request, 'vacay/edit_employee.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_employee(request, employee_id):
    if request.method == 'POST':

        name = request.POST.get('name', None)
        gender = request.POST.get('gender', None)
        email = request.POST.get('email', None)
        password = request.POST.get('password', None)
        millennial = request.POST.get('millennial', None)
        bucks = request.POST.get('bucks', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        employee = Employee.objects.get(em_id=employee_id)

        employee.em_name = name
        employee.em_gender = gender
        employee.em_email = email
        employee.em_password = password
        employee.em_millennial = millennial
        if "$" in str(bucks):
            employee.em_givenbuck = str(bucks)
        else:
            employee.em_givenbuck = "$"+ str(bucks)

        try:
            image = request.FILES['photo']
            fs = FileSystemStorage()
            try:
                x = request.POST.get('x', '0')
                y = request.POST.get('y', '0')
                w = request.POST.get('w', '32')
                h = request.POST.get('h', '32')
                #  return HttpResponse(w)
                file = profile_process(image, x, y, w, h)
                image = file
                # return HttpResponse('Cropped!')

            except MultiValueDictKeyError:
                print('No cropping')
            except ValueError:
                print('No cropping')

            filename = fs.save(image.name, image)
            uploaded_file_url = fs.url(filename)
            employee.em_image = settings.URL + uploaded_file_url

        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        employee.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        employees = Employee.objects.filter(adminID=idx).order_by('-id')
        context = {'employees': employees, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_employees.html', context)

    elif request.method == 'GET':
        pass

def delete_employee(request, employee_id):

    Employee.objects.get(em_id=employee_id).delete()

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    employees = Employee.objects.filter(adminID=idx).order_by('-id')
    context = {'employees': employees, 'admin': adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_employees.html', context)

def add_employee(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    return render(request, 'vacay/add_employee.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_employee_process(request):
    if request.method == 'POST':

        name = request.POST.get('name', None)
        gender = request.POST.get('gender', None)
        email = request.POST.get('email', None)
        password = request.POST.get('password', None)
        millennial = request.POST.get('millennial', None)
        bucks = request.POST.get('bucks', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        employee = Employee()
        employee.adminID = idx
        employee.em_name = name
        employee.em_gender = gender
        employee.em_email = email
        employee.em_password = password
        employee.em_millennial = millennial
        if "$" in str(bucks):
            employee.em_givenbuck = str(bucks)
        else:
            employee.em_givenbuck = "$"+ str(bucks)

        image = request.FILES['photo']
        fs = FileSystemStorage()
        try:
            x = request.POST.get('x', '0')
            y = request.POST.get('y', '0')
            w = request.POST.get('w', '32')
            h = request.POST.get('h', '32')
            #  return HttpResponse(w)
            file = profile_process(image, x, y, w, h)
            image = file
            # return HttpResponse('Cropped!')

        except MultiValueDictKeyError:
            print('No cropping')
        except ValueError:
            print('No cropping')

        filename = fs.save(image.name, image)
        uploaded_file_url = fs.url(filename)
        employee.em_image = settings.URL + uploaded_file_url

        employee.save()
        employee.em_id = employee.pk
        employee.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        employees = Employee.objects.filter(adminID=idx).order_by('-id')
        context = {'employees': employees, 'admin':adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_employees.html', context)

    elif request.method == 'GET':
        pass

def show_jobs(request):
    global com
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    com=0
    request.session['com']=0
    adminUser = AdminUser.objects.get(adminID=idx)
    jobs=Job.objects.filter(adminID=idx).order_by('-id')
    context = {'jobs': jobs, 'admin':adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_jobs.html', context)

    # return HttpResponse(idx)

def add_job(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    return render(request, 'vacay/add_job.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_job_process(request):
    if request.method == 'POST':

        jobtitle = request.POST.get('jobtitle', None)
        reqid = request.POST.get('reqid', None)
        department = request.POST.get('department', None)
        location = request.POST.get('location', None)
        # postingdate = request.POST.get('postingdate', None)
        description = request.POST.get('description', None)
        extra = request.POST.get('extra', None)
        survey_link = request.POST.get('survey', None)
        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        date = strftime("%Y-%m-%d", gmtime())

        job = Job()
        job.adminID = idx
        job.job_name = jobtitle
        job.job_req = reqid
        job.job_department = department
        job.job_location = location
        job.job_postdate = date
        job.job_description = description
        job.job_empty = extra
        job.job_survey = survey_link

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            job.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            job.youtube_url = youtubeurl

        job.save()
        job.job_id = job.pk
        job.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        jobs = Job.objects.filter(adminID=idx).order_by('-id')
        context = {'jobs': jobs, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_jobs.html', context)

    elif request.method == 'GET':
        pass

def edit_job(request, job_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    job=Job.objects.get(job_id=job_id)
    context = {'job': job}
    # cache.set('search_env', 3)
    return render(request, 'vacay/edit_job.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_job(request, job_id):
    if request.method == 'POST':

        jobtitle = request.POST.get('jobtitle', None)
        reqid = request.POST.get('reqid', None)
        department = request.POST.get('department', None)
        location = request.POST.get('location', None)
        # postingdate = request.POST.get('postingdate', None)
        description = request.POST.get('description', None)
        extra = request.POST.get('extra', None)
        survey_link = request.POST.get('survey', None)

        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        date = strftime("%Y-%m-%d", gmtime())

        job = Job.objects.get(job_id=job_id)

        job.job_name = jobtitle
        job.job_req = reqid
        job.job_department = department
        job.job_location = location
        job.job_postdate = "Updated at: "+date
        job.job_description = description
        job.job_empty = extra
        job.job_survey = survey_link

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            job.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            job.youtube_url = youtubeurl

        job.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        jobs = Job.objects.filter(adminID=idx).order_by('-id')
        context = {'jobs': jobs, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_jobs.html', context)

    elif request.method == 'GET':
        pass

def delete_job(request, job_id):

    Job.objects.get(job_id=job_id).delete()

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    jobs = Job.objects.filter(adminID=idx).order_by('-id')
    context = {'jobs': jobs, 'admin': adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_jobs.html', context)

def show_announcements(request):
    global com
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    com=0
    request.session['com']=0
    adminUser = AdminUser.objects.get(adminID=idx)
    announces=Announce.objects.filter(adminID=idx).order_by('-id')
    context = {'announces': announces, 'admin': adminUser}
    return render(request, 'vacay/show_announcements.html', context)

def edit_announcement(request, announce_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    announce=Announce.objects.get(an_id=announce_id)
    context = {'announce': announce}
    # cache.set('search_env', 3)
    return render(request, 'vacay/edit_announcement.html', context)

def add_announcement(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    return render(request, 'vacay/add_announcement.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_announcement_process(request):
    if request.method == 'POST':

        announcementtitle = request.POST.get('announcementtitle', None)
        audience = request.POST.get('audience', None)
        subject = request.POST.get('subject', None)
        callofaction = request.POST.get('callofaction', None)
        owneremail = request.POST.get('owneremail', None)
        description = request.POST.get('description', None)
        survey = request.POST.get('survey', None)
        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        date=strftime("%Y-%m-%d", gmtime())

        announce = Announce()
        announce.adminID = idx
        announce.an_title = announcementtitle
        announce.an_audience = audience
        announce.an_subject = subject
        announce.an_description = description
        announce.an_callofaction = callofaction
        announce.an_owneremail = owneremail
        announce.an_postdate = date
        announce.an_survey = survey

        try:
            image = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            announce.an_image = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            announce.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            announce.youtube_url = youtubeurl

        announce.save()
        announce.an_id = announce.pk
        announce.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        announces = Announce.objects.filter(adminID=idx).order_by('-id')
        context = {'announces': announces, 'admin':adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_announcements.html', context)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_announcement(request, announce_id):

    if request.method == 'POST':

        announcementtitle = request.POST.get('announcementtitle', None)
        audience = request.POST.get('audience', None)
        subject = request.POST.get('subject', None)
        callofaction = request.POST.get('callofaction', None)
        owneremail = request.POST.get('owneremail', None)
        description = request.POST.get('description', None)
        survey = request.POST.get('survey', None)
        youtubeurl = request.POST.get('youtubeurl', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        date = strftime("%Y-%m-%d", gmtime())

        announce = Announce.objects.get(an_id=announce_id)

        announce.an_title = announcementtitle
        announce.an_audience = audience
        announce.an_subject = subject
        announce.an_callofaction = callofaction
        announce.an_owneremail = owneremail
        announce.an_description = description
        announce.an_survey = survey
        announce.an_postdate = "Updated at: "+date

        try:
            image = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            announce.an_image = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            announce.video_url = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            announce.youtube_url = youtubeurl

        announce.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        announces = Announce.objects.filter(adminID=idx).order_by('-id')
        context = {'announces': announces, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_announcements.html', context)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updateannouncementpictures(request, announce_id):

    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        date = strftime("%Y-%m-%d", gmtime())

        announce = Announce.objects.get(an_id=announce_id)

        try:
            image = request.FILES['photoa']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            announce.imageA = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photob']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            announce.imageB = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photoc']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            announce.imageC = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photod']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            announce.imageD = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photoe']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            announce.imageE = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photof']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            announce.imageF = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        desca = request.POST.get('desca', '')
        descb = request.POST.get('descb', '')
        descc = request.POST.get('descc', '')
        descd = request.POST.get('descd', '')
        desce = request.POST.get('desce', '')
        descf = request.POST.get('descf', '')

        announce.descA=desca
        announce.descB=descb
        announce.descC=descc
        announce.descD=descd
        announce.descE=desce
        announce.descF=descf

        announce.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        announces = Announce.objects.filter(adminID=idx).order_by('-id')
        context = {'announces': announces, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_announcements.html', context)

    elif request.method == 'GET':
        pass

def delete_announcement(request, announce_id):

    Announce.objects.get(an_id=announce_id).delete()

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    announces = Announce.objects.filter(adminID=idx).order_by('-id')
    context = {'announces': announces, 'admin': adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_announcements.html', context)

@csrf_protect
@csrf_exempt
def search_provider(request):
    global serv2
    global prod
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=idx)
        search_id = request.POST.get('q', None)

        serv2=request.session['serv2']
        prod=request.session['prod']

        providers = Provider.objects.filter(proFirstName__contains=search_id, adminID=idx)
        if providers.count()>0:
            if serv2 == 1:
                context = {'all_providers': providers, 'admin': adminUser, 'note': 'service'}
                return render(request, 'vacay/select_provider.html', context)
            elif prod == 1:
                context = {'all_providers': providers, 'admin': adminUser, 'note': 'product'}
                return render(request, 'vacay/select_provider.html', context)
            else:
                context = {'all_providers': providers, 'admin': adminUser}
                return render(request, 'vacay/index.html', context)

        else:
            providers = Provider.objects.filter(proLastName__contains=search_id, adminID=idx)
            if providers.count() > 0:
                if serv2==1:
                    context = {'all_providers': providers, 'admin': adminUser, 'note':'service'}
                    return render(request, 'vacay/select_provider.html', context)
                elif prod==1:
                    context = {'all_providers': providers, 'admin': adminUser, 'note':'product'}
                    return render(request, 'vacay/select_provider.html', context)
                else:
                    context = {'all_providers': providers, 'admin': adminUser}
                    return render(request, 'vacay/index.html', context)
            else:
                providers = Provider.objects.filter(proEmail__contains=search_id, adminID=idx)
                if providers.count() > 0:
                    if serv2 == 1:
                        context = {'all_providers': providers, 'admin': adminUser, 'note': 'service'}
                        return render(request, 'vacay/select_provider.html', context)
                    elif prod == 1:
                        context = {'all_providers': providers, 'admin': adminUser, 'note': 'product'}
                        return render(request, 'vacay/select_provider.html', context)
                    else:
                        context = {'all_providers': providers, 'admin': adminUser}
                        return render(request, 'vacay/index.html', context)
                else:
                    providers = Provider.objects.filter(proCity__contains=search_id, adminID=idx)
                    if providers.count() > 0:
                        if serv2 == 1:
                            context = {'all_providers': providers, 'admin': adminUser, 'note': 'service'}
                            return render(request, 'vacay/select_provider.html', context)
                        elif prod == 1:
                            context = {'all_providers': providers, 'admin': adminUser, 'note': 'product'}
                            return render(request, 'vacay/select_provider.html', context)
                        else:
                            context = {'all_providers': providers, 'admin': adminUser}
                            return render(request, 'vacay/index.html', context)
                    else:
                        providers = Provider.objects.filter(proCompany__contains=search_id, adminID=idx)
                        if serv2 == 1:
                            context = {'all_providers': providers, 'admin': adminUser, 'note': 'service'}
                            return render(request, 'vacay/select_provider.html', context)
                        elif prod == 1:
                            context = {'all_providers': providers, 'admin': adminUser, 'note': 'product'}
                            return render(request, 'vacay/select_provider.html', context)
                        else:
                            context = {'all_providers': providers, 'admin': adminUser}
                            return render(request, 'vacay/index.html', context)

    else:
        pass

@csrf_protect
@csrf_exempt
def search_service(request):

    global pro_num
    global serv2
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=idx)
        search_id = request.POST.get('q', None)

        pro_num=request.session['pro_id']
        serv2=request.session['serv2']

        provider = Provider.objects.get(proid=pro_num)
        services = Service.objects.filter(proBeautyCategory__contains=search_id, adminID=idx)
        if services.count()>0:
            context = {'services': services, 'provider': provider}
            if serv2==1:
                return render(request, 'vacay/service_list.html', context)
            else:
                return render(request, 'vacay/show_services.html', context)
        else:
            services = Service.objects.filter(proBeautySubCategory__contains=search_id, adminID=idx)
            if services.count()>0:
                context = {'services': services, 'provider': provider}
                if serv2 == 1:
                    return render(request, 'vacay/service_list.html', context)
                else:
                    return render(request, 'vacay/show_services.html', context)
            else:
                services = Service.objects.filter(proServicePrice__contains=search_id, adminID=idx)
                context = {'services': services, 'provider': provider}
                if serv2 == 1:
                    return render(request, 'vacay/service_list.html', context)
                else:
                    return render(request, 'vacay/show_services.html', context)
    else:
        pass

@csrf_protect
@csrf_exempt
def search_product(request):

    global pro_num
    global prod
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        search_id = request.POST.get('q', None)

        pro_num=request.session['pro_id']
        prod=request.session['prod']

        provider = Provider.objects.get(proid=pro_num)
        product = Product.objects.filter(itemBrand__contains=search_id, proid=pro_num)
        if product.count()>0:
            context = {'product': product, 'provider': provider}
            if prod==1:
                return render(request, 'vacay/product_list.html', context)
            else:
                return render(request, 'vacay/show_products.html', context)
        else:
            product = Product.objects.filter(itemProduct__contains=search_id, proid=pro_num)
            if product.count()>0:
                context = {'product': product, 'provider': provider}
                if prod == 1:
                    return render(request, 'vacay/product_list.html', context)
                else:
                    return render(request, 'vacay/show_products.html', context)
            else:
                product = Product.objects.filter(itemName__contains=search_id, proid=pro_num)
                if product.count()>0:
                    context = {'product': product, 'provider': provider}
                    if prod == 1:
                        return render(request, 'vacay/product_list.html', context)
                    else:
                        return render(request, 'vacay/show_products.html', context)
                else:
                    product = Product.objects.filter(itemPrice__contains=search_id, proid=pro_num)
                    if product.count() > 0:
                        context = {'product': product, 'provider': provider}
                        if prod == 1:
                            return render(request, 'vacay/product_list.html', context)
                        else:
                            return render(request, 'vacay/show_products.html', context)
                    else:
                        product = Product.objects.filter(itemSize__contains=search_id, proid=pro_num)
                        context = {'product': product, 'provider': provider}
                        if prod == 1:
                            return render(request, 'vacay/product_list.html', context)
                        else:
                            return render(request, 'vacay/show_products.html', context)
    else:
        pass


@csrf_protect
@csrf_exempt
def search_provider_setup(request):
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=idx)
        search_id = request.POST.get('q', None)

        provider = Provider.objects.filter(proFirstName__contains=search_id, adminID=idx)
        if provider.count()>0:
            context = {'provider': provider, 'admin': adminUser}
            # cache.set('search_env', 3)
            return render(request, 'vacay/show_provider_setup.html', context)

        else:
            provider = Provider.objects.filter(proLastName__contains=search_id, adminID=idx)
            if provider.count() > 0:
                context = {'provider': provider, 'admin': adminUser}
                return render(request, 'vacay/show_provider_setup.html', context)
            else:
                provider = Provider.objects.filter(proServicePercent__contains=search_id, adminID=idx)
                if provider.count() > 0:
                    context = {'provider': provider, 'admin': adminUser}
                    # cache.set('search_env', 3)
                    return render(request, 'vacay/show_provider_setup.html', context)
                else:
                    provider = Provider.objects.filter(proSalary__contains=search_id, adminID=idx)
                    if provider.count() > 0:
                        context = {'provider': provider, 'admin': adminUser}
                        # cache.set('search_env', 3)
                        return render(request, 'vacay/show_provider_setup.html', context)
                    else:
                        provider = Provider.objects.filter(proProductSalePercent__contains=search_id, adminID=idx)
                        context = {'provider': provider, 'admin': adminUser}
                        # cache.set('search_env', 3)
                        return render(request, 'vacay/show_provider_setup.html', context)

    else:
        pass

@csrf_protect
@csrf_exempt
def search_job(request):
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=idx)
        search_id = request.POST.get('q', None)

        jobs = Job.objects.filter(job_name__contains=search_id, adminID=idx)
        if jobs.count()>0:
            context = {'jobs': jobs, 'admin': adminUser}
            return render(request, 'vacay/show_jobs.html', context)

        else:
            jobs = Job.objects.filter(job_req__contains=search_id, adminID=idx)
            if jobs.count() > 0:
                context = {'jobs': jobs, 'admin': adminUser}
                return render(request, 'vacay/show_jobs.html', context)
            else:
                jobs = Job.objects.filter(job_department__contains=search_id, adminID=idx)
                if jobs.count() > 0:
                    context = {'jobs': jobs, 'admin': adminUser}
                    return render(request, 'vacay/show_jobs.html', context)
                else:
                    jobs = Job.objects.filter(job_location__contains=search_id, adminID=idx)
                    if jobs.count() > 0:
                        context = {'jobs': jobs, 'admin': adminUser}
                        return render(request, 'vacay/show_jobs.html', context)
                    else:
                        jobs = Job.objects.filter(job_postdate__contains=search_id, adminID=idx)
                        context = {'jobs': jobs, 'admin': adminUser}
                        return render(request, 'vacay/show_jobs.html', context)

    else:
        pass

@csrf_protect
@csrf_exempt
def search_employee(request):
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=idx)
        search_id = request.POST.get('q', None)

        employees = Employee.objects.filter(em_name__contains=search_id, adminID=idx)
        if employees.count()>0:
            context = {'employees': employees, 'admin': adminUser}
            return render(request, 'vacay/show_employees.html', context)

        else:
            employees = Employee.objects.filter(em_gender__contains=search_id, adminID=idx)
            if employees.count() > 0:
                context = {'employees': employees, 'admin': adminUser}
                return render(request, 'vacay/show_employees.html', context)
            else:
                employees = Employee.objects.filter(em_email__contains=search_id, adminID=idx)
                if employees.count() > 0:
                    context = {'employees': employees, 'admin': adminUser}
                    return render(request, 'vacay/show_employees.html', context)
                else:
                    employees = Employee.objects.filter(em_millennial__contains=search_id, adminID=idx)
                    if employees.count() > 0:
                        context = {'employees': employees, 'admin': adminUser}
                        return render(request, 'vacay/show_employees.html', context)
                    else:
                        employees = Employee.objects.filter(em_interaction__contains=search_id, adminID=idx)
                        if employees.count() > 0:
                            context = {'employees': employees, 'admin': adminUser}
                            return render(request, 'vacay/show_employees.html', context)
                        else:
                            employees = Employee.objects.filter(em_givenbuck__contains=search_id, adminID=idx)
                            if employees.count() > 0:
                                context = {'employees': employees, 'admin': adminUser}
                                return render(request, 'vacay/show_employees.html', context)
                            else:
                                employees = Employee.objects.filter(em_usedbuck__contains=search_id, adminID=idx)
                                context = {'employees': employees, 'admin': adminUser}
                                return render(request, 'vacay/show_employees.html', context)

    else:
        pass

@csrf_protect
@csrf_exempt
def search_announce(request):
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=idx)
        search_id = request.POST.get('q', None)

        announces = Announce.objects.filter(an_title__contains=search_id, adminID=idx)
        if announces.count()>0:
            context = {'announces': announces, 'admin': adminUser}
            return render(request, 'vacay/show_announcements.html', context)

        else:
            announces = Announce.objects.filter(an_audience__contains=search_id, adminID=idx)
            if announces.count() > 0:
                context = {'announces': announces, 'admin': adminUser}
                return render(request, 'vacay/show_announcements.html', context)
            else:
                announces = Announce.objects.filter(an_subject__contains=search_id, adminID=idx)
                if announces.count() > 0:
                    context = {'announces': announces, 'admin': adminUser}
                    return render(request, 'vacay/show_announcements.html', context)
                else:
                    announces = Announce.objects.filter(an_owneremail__contains=search_id, adminID=idx)
                    if announces.count() > 0:
                        context = {'announces': announces, 'admin': adminUser}
                        return render(request, 'vacay/show_announcements.html', context)
                    else:
                        announces = Announce.objects.filter(an_viewnum__contains=search_id, adminID=idx)
                        if announces.count() > 0:
                            context = {'announces': announces, 'admin': adminUser}
                            return render(request, 'vacay/show_announcements.html', context)
                        else:
                            announces = Announce.objects.filter(an_responsenum__contains=search_id, adminID=idx)
                            if announces.count() > 0:
                                context = {'announces': announces, 'admin': adminUser}
                                return render(request, 'vacay/show_announcements.html', context)
                            else:
                                announces = Announce.objects.filter(an_callofaction__contains=search_id, adminID=idx)
                                if announces.count() > 0:
                                    context = {'announces': announces, 'admin': adminUser}
                                    return render(request, 'vacay/show_announcements.html', context)
                                else:
                                    announces = Announce.objects.filter(an_postdate__contains=search_id, adminID=idx)
                                    context = {'announces': announces, 'admin': adminUser}
                                    return render(request, 'vacay/show_announcements.html', context)

    else:
        pass

@csrf_protect
@csrf_exempt
def search_broadmoor_product(request):
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=idx)
        search_id = request.POST.get('q', None)

        bproducts = BroadmoorProduct.objects.filter(bm_proName__contains=search_id, adminID=idx)
        if bproducts.count()>0:
            context = {'bproducts': bproducts, 'admin': adminUser}
            return render(request, 'vacay/show_broadmoor_products.html', context)

        else:
            bproducts = BroadmoorProduct.objects.filter(bm_proInventoryNum__contains=search_id, adminID=idx)
            if bproducts.count() > 0:
                context = {'bproducts': bproducts, 'admin': adminUser}
                return render(request, 'vacay/show_broadmoor_products.html', context)
            else:
                bproducts = BroadmoorProduct.objects.filter(bm_proCategory__contains=search_id, adminID=idx)
                context = {'bproducts': bproducts, 'admin': adminUser}
                return render(request, 'vacay/show_broadmoor_products.html', context)

    else:
        pass


def export_xlsx_provider(request):
    import openpyxl
    from openpyxl.utils import get_column_letter

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=provider_template.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Providers"

    row_num = 0

    columns = [
        (u"Picture Url", 40),
        (u"First Name", 15),
        (u"Last Name", 15),
        (u"Email", 40),
        (u"Password", 30),
        (u"Phone Number", 30),
        (u"City (from Sheet2)", 20),
        (u"Address", 30),
        (u"Company", 20),
        (u"Percent of Service", 25),
        (u"Salary Amount", 20),
        (u"Percent of Product Sales", 30),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    ws2=wb.create_sheet(title='Sheet2')

    ws2.column_dimensions["A"].width = 20
    my_color = openpyxl.styles.colors.Color(rgb='00ffaa02')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    ws2['A1'].fill = my_fill
    ws2['A1'] = 'City'
    ws2['A2'] = 'San Francisco'
    ws2['A3'] = 'New York'
    ws2['A4'] = 'Chicago'
    ws2['A5'] = 'Denver'

    wb.save(response)
    return response


def import_view_provider(request):

    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()
    return render(
        request,
        'vacay/upload_form.html',
        {
            'form': form,
            'title': 'Load Data',
            'header': 'Upload provider information:'
        })

def import_provider_data(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        if form.is_valid():

            if not request.session.exists(request.session.session_key):
                request.session.create()
            session_key = request.session.session_key
            session = Session.objects.get(session_key=session_key)
            idx = session.get_decoded().get('_auth_user_id')

            if idx is None:
                return redirect('/logout')

            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    pictureUrl = sheet.cell(r, 0).value
                    firstName = sheet.cell(r, 1).value
                    lastName = sheet.cell(r, 2).value
                    email = sheet.cell(r, 3).value
                    password = sheet.cell(r, 4).value
                    phone = sheet.cell(r, 5).value
                    city = sheet.cell(r, 6).value
                    address = sheet.cell(r, 7).value
                    company = sheet.cell(r, 8).value
                    servicepercent = sheet.cell(r, 9).value
                    salaryamount = sheet.cell(r, 10).value
                    productpercent = sheet.cell(r, 11).value

                    provider = Provider()
                    provider.adminID = idx
                    provider.proFirstName = firstName
                    provider.proLastName = lastName
                    provider.proEmail = email
                    provider.proPassword = password
                    provider.proPhone = phone
                    provider.proCity = city
                    provider.proAddress = address
                    provider.proCompany = company
                    provider.proServicePercent = servicepercent
                    provider.proSalary = str(salaryamount).replace(".0","")
                    provider.proProductSalePercent = productpercent
                    provider.proProfileImageUrl = pictureUrl

                    provider.save()
                    provider.proid = provider.pk
                    provider.save()

                adminUser = AdminUser.objects.get(adminID=idx)
                all_providers = Provider.objects.filter(adminID=idx).order_by('-id')
                context = {'all_providers': all_providers, 'admin': adminUser}
                # cache.set('search_env', 3)
                return render(request, 'vacay/index.html', context)
            except XLRDError:
                return render(request, 'vacay/upload_form.html', {'note':'invalid_file'})
            except IOError:
                return render(request, 'vacay/upload_form.html', {'note': 'invalid_file'})
            except IndexError:
                return render(request, 'vacay/upload_form_service.html', {'note': 'invalid_file'})
            except DataError:
                return render(request, 'vacay/upload_form_service.html', {'note': 'invalid_file'})
        else:
            return render(request, 'vacay/upload_form.html', {'note':'invalid_file'})

    elif request.method == 'GET':
        pass

def export_xlsx_service(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=service_template.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Services"

    row_num = 0

    columns = [
        (u"Picture Url", 40),
        (u"Beauty Category (from Sheet2)", 30),
        (u"Service Name (from Sheet2)", 30),
        (u"Price", 20),
        (u"Description", 40),
        (u"Video URL(Optional)", 40),
        (u"YouTube Video ID(Optional)", 40),
        (u"Extra Picture URL A(Optional)", 40),
        (u"Extra Picture URL B(Optional)", 40),
        (u"Extra Picture URL C(Optional)", 40),
        (u"Extra Picture URL D(Optional)", 40),
        (u"Extra Picture URL E(Optional)", 40),
        (u"Extra Picture URL F(Optional)", 40),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    ws2 = wb.create_sheet(title='Sheet2')

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 40
    my_color1 = openpyxl.styles.colors.Color(rgb='00ffaa02')
    my_fill1 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color1)
    my_color2 = openpyxl.styles.colors.Color(rgb='0000e5ff')
    my_fill2 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color2)

    ws2['A1'].fill = my_fill1
    ws2["A1"] = 'Service Category'
    ws2["A2"] = 'Hair(Women)'
    ws2["A7"] = 'Blowout'
    ws2["A8"] = 'Manicure/Pedicure(Women)'
    ws2["A13"] = 'Massage(Women)'
    ws2["A17"] = 'Wax(Women)'
    ws2["A21"] = 'Facial(Women)'
    ws2["A23"] = 'Makeover'
    ws2["A24"] = 'Hair(Men)'
    ws2["A28"] = 'Manicure/Pedicure(Men)'
    ws2["A31"] = 'Hot Shave'
    ws2["A35"] = 'Wax(Men)'
    ws2["A39"] = 'Facial(Men)'
    ws2["A43"] = 'Massage(Men)'

    ws2['B1'].fill = my_fill2
    ws2["B1"] = 'Service Name'
    ws2["B2"] = 'Haircut'
    ws2["B3"] = 'Color'
    ws2["B4"] = 'Brazilian Blowout'
    ws2["B5"] = 'Keratin Treatment'
    ws2["B6"] = 'Deep Conditioner'
    ws2["B7"] = 'Blowout'
    ws2["B8"] = 'Manicure'
    ws2["B9"] = 'Manicure: Gel'
    ws2["B10"] = 'Pedicure'
    ws2["B11"] = 'Pedicure: Gel'
    ws2["B12"] = 'Pink & White'
    ws2["B13"] = 'Deep Tissue Massage(50 minutes)'
    ws2["B14"] = 'Deep Tissue Massage(90 minutes)'
    ws2["B15"] = 'Swedish Massage(50 minutes)'
    ws2["B16"] = 'Swedish Massage(90 minutes)'
    ws2["B17"] = 'Eye Brow Wax'
    ws2["B18"] = 'Lip Wax'
    ws2["B19"] = 'Bikini Wax'
    ws2["B20"] = 'Brazilian Wax'
    ws2["B21"] = 'Basic Facial'
    ws2["B22"] = 'Premium Facial'
    ws2["B23"] = 'Makeover: Seasonal Trends'
    ws2["B24"] = 'Classic Cut'
    ws2["B25"] = 'Frosty Color'
    ws2["B26"] = 'Trendy Cut'
    ws2["B27"] = 'Classic Color'
    ws2["B28"] = 'MANicure'
    ws2["B29"] = 'PIECE OF HEAVEN CURE'
    ws2["B30"] = 'Feet Meet Your Never-ending Treat'
    ws2["B31"] = 'Executive Shave'
    ws2["B32"] = 'Chairman of the Board Shave'
    ws2["B33"] = 'Neck Cleanup'
    ws2["B34"] = 'Hot Towel Shave'
    ws2["B35"] = 'Chest Wax'
    ws2["B36"] = 'Nose Wax'
    ws2["B37"] = 'Back Wax'
    ws2["B38"] = 'Eyebrow Wax'
    ws2["B39"] = 'Rejuvenating Eye Masque'
    ws2["B40"] = 'Fountain of Youth Masque'
    ws2["B41"] = 'De-Stress Facial'
    ws2["B42"] = 'Gentleman\'s Facial'
    ws2["B43"] = 'Deep Tissue Massage'
    ws2["B44"] = 'Swedish Massage'
    ws2["B45"] = 'Neck & Back Massage'
    ws2["B46"] = 'Sugar Massage'

    wb.save(response)
    return response

def import_view_service(request):
    global serv2

    serv2 = request.session['serv2']

    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()

    if serv2==1:
        return render(
            request,
            'vacay/upload_form_service_home.html',
            {
                'form': form,
                'title': 'Load Data',
                'header': 'Upload beauty service info:'
            })
    else:
        return render(
            request,
            'vacay/upload_form_service.html',
            {
                'form': form,
                'title': 'Load Data',
                'header': 'Upload beauty service info:'
            })

def import_service_data(request):

    global pro_num
    global serv2

    serv2 = request.session['serv2']
    pro_num=request.session['pro_id']

    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        if form.is_valid():

            if not request.session.exists(request.session.session_key):
                request.session.create()
            session_key = request.session.session_key
            session = Session.objects.get(session_key=session_key)
            idx = session.get_decoded().get('_auth_user_id')

            if idx is None:
                return redirect('/logout')

            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    pictureUrl = sheet.cell(r, 0).value
                    category = sheet.cell(r, 1).value
                    serviceName = sheet.cell(r, 2).value
                    price = sheet.cell(r, 3).value
                    description = sheet.cell(r, 4).value

                    video = sheet.cell(r, 5).value
                    youtube = sheet.cell(r, 6).value
                    imageA = sheet.cell(r, 7).value
                    imageB = sheet.cell(r, 8).value
                    imageC = sheet.cell(r, 9).value
                    imageD = sheet.cell(r, 10).value
                    imageE = sheet.cell(r, 11).value
                    imageF = sheet.cell(r, 12).value

                    service = Service()
                    service.adminID = idx
                    service.proid = pro_num
                    service.proBeautyCategory = category
                    service.proBeautySubCategory = serviceName

                    if "$" in str(price):
                        service.proServicePrice = str(price)
                    else:
                        service.proServicePrice = "$" + str(price)

                    service.proServiceDescription = description
                    service.proServicePictureUrl = pictureUrl

                    service.video_url=video
                    service.youtube_url=youtube
                    service.imageA=imageA
                    service.imageB=imageB
                    service.imageC=imageC
                    service.imageD=imageD
                    service.imageE=imageE
                    service.imageF=imageF

                    service.save()
                    service.serviceid = service.pk
                    service.save()

                provider = Provider.objects.get(proid=pro_num)
                services = Service.objects.filter(proid=pro_num).order_by('-id')
                context = {'services': services, 'provider': provider}
                # cache.set('search_env', 3)
                if serv2==1:
                    return render(request, 'vacay/service_list.html', context)
                else:
                    return render(request, 'vacay/show_services.html', context)

            except XLRDError:
                if serv2==1:
                    return render(request, 'vacay/upload_form_service_home.html', {'note': 'invalid_file'})
                else:
                    return render(request, 'vacay/upload_form_service.html', {'note': 'invalid_file'})

            except IOError:
                if serv2==1:
                    return render(request, 'vacay/upload_form_service_home.html', {'note': 'invalid_file'})
                else:
                    return render(request, 'vacay/upload_form_service.html', {'note': 'invalid_file'})
            except IndexError:
                if serv2==1:
                    return render(request, 'vacay/upload_form_service_home.html', {'note': 'invalid_file'})
                else:
                    return render(request, 'vacay/upload_form_service.html', {'note': 'invalid_file'})
            except DataError:
                return HttpResponse('Invalid file!')
        else:
            if serv2 == 1:
                return render(request, 'vacay/upload_form_service_home.html', {'note': 'invalid_file'})
            else:
                return render(request, 'vacay/upload_form_service.html', {'note': 'invalid_file'})

    elif request.method == 'GET':
        pass

def export_xlsx_product(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=product_template.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Products"

    row_num = 0

    columns = [
        (u"Picture Url", 40),
        (u"Brand", 30),
        (u"Product", 30),
        (u"Product Name", 30),
        (u"Size", 30),
        (u"Price", 20),
        (u"Inventory#", 30),
        (u"Sale Status(0:Not Sold/1:Sold)", 40),
        (u"Description", 40),
        (u"Video URL(Optional)", 40),
        (u"YouTube Video ID(Optional)", 40),
        (u"Extra Picture URL A(Optional)", 40),
        (u"Extra Picture URL B(Optional)", 40),
        (u"Extra Picture URL C(Optional)", 40),
        (u"Extra Picture URL D(Optional)", 40),
        (u"Extra Picture URL E(Optional)", 40),
        (u"Extra Picture URL F(Optional)", 40),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    wb.save(response)
    return response

def import_view_product(request):
    global prod
    prod = request.session['prod']
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()

    if prod == 1:
        return render(
            request,
            'vacay/upload_form_product_home.html',
            {
                'form': form,
                'title': 'Load Data',
                'header': 'Upload beauty product info:'
            })
    else:
        return render(
            request,
            'vacay/upload_form_product.html',
            {
                'form': form,
                'title': 'Load Data',
                'header': 'Upload beauty product info:'
            })

def import_product_data(request):

    global pro_num
    global prod
    pro_num = request.session['pro_id']
    prod = request.session['prod']
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        if form.is_valid():

            if not request.session.exists(request.session.session_key):
                request.session.create()
            session_key = request.session.session_key
            session = Session.objects.get(session_key=session_key)
            idx = session.get_decoded().get('_auth_user_id')

            if idx is None:
                return redirect('/logout')

            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    pictureUrl = sheet.cell(r, 0).value
                    brand = sheet.cell(r, 1).value
                    product_t = sheet.cell(r, 2).value
                    productName = sheet.cell(r, 3).value
                    size = sheet.cell(r, 4).value
                    price = sheet.cell(r, 5).value
                    inventory = sheet.cell(r, 6).value
                    saleStatus = sheet.cell(r, 7).value
                    description = sheet.cell(r, 8).value

                    video = sheet.cell(r, 9).value
                    youtube = sheet.cell(r, 10).value
                    imageA = sheet.cell(r, 11).value
                    imageB = sheet.cell(r, 12).value
                    imageC = sheet.cell(r, 13).value
                    imageD = sheet.cell(r, 14).value
                    imageE = sheet.cell(r, 15).value
                    imageF = sheet.cell(r, 16).value

                    product = Product()
                    product.proid = pro_num
                    product.itemProduct = product_t
                    product.itemName = productName
                    product.itemBrand = brand
                    product.itemSize = size

                    if "$" in str(price):
                        product.itemPrice = str(price)
                    else:
                        product.itemPrice = "$"+str(price)

                    product.itemInventoryNum = str(inventory).replace(".0","")
                    product.itemDescription = description
                    product.itemSaleStatus = saleStatus
                    product.itemPictureUrl = pictureUrl

                    product.video_url = video
                    product.youtube_url = youtube
                    product.imageA = imageA
                    product.imageB = imageB
                    product.imageC = imageC
                    product.imageD = imageD
                    product.imageE = imageE
                    product.imageF = imageF

                    product.save()
                    product.itemid = product.pk
                    product.save()

                provider = Provider.objects.get(proid=pro_num)
                product = Product.objects.filter(proid=pro_num).order_by('-id')
                context = {'product': product, 'provider': provider}
                if prod==1:
                    return render(request, 'vacay/product_list.html', context)
                else:
                    return render(request, 'vacay/show_products.html', context)

            except XLRDError:
                if prod==1:
                    return render(request, 'vacay/upload_form_product_home.html', {'note': 'invalid_file'})
                else:
                    return render(request, 'vacay/upload_form_product.html', {'note': 'invalid_file'})

            except IOError:
                if prod==1:
                    return render(request, 'vacay/upload_form_product_home.html', {'note': 'invalid_file'})
                else:
                    return render(request, 'vacay/upload_form_product.html', {'note': 'invalid_file'})
            except IndexError:
                if prod==1:
                    return render(request, 'vacay/upload_form_product_home.html', {'note': 'invalid_file'})
                else:
                    return render(request, 'vacay/upload_form_product.html', {'note': 'invalid_file'})
            except DataError:
                return HttpResponse('Invalid file!')
        else:
            if prod == 1:
                return render(request, 'vacay/upload_form_product_home.html', {'note': 'invalid_file'})
            else:
                return render(request, 'vacay/upload_form_product.html', {'note': 'invalid_file'})

    elif request.method == 'GET':
        pass

def export_xlsx_broadmoor(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=retail_template.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Retail"

    row_num = 0

    columns = [
        (u"Picture Url", 40),
        (u"Product Name", 30),
        (u"Inventory#", 30),
        (u"Category (from Sheet2)", 30),
        (u"Additional Marketing Material", 40),
        (u"Video URL(Optional)", 40),
        (u"YouTube Video ID(Optional)", 40),
        (u"Extra Picture URL A(Optional)", 40),
        (u"Extra Picture URL B(Optional)", 40),
        (u"Extra Picture URL C(Optional)", 40),
        (u"Extra Picture URL D(Optional)", 40),
        (u"Extra Picture URL E(Optional)", 40),
        (u"Extra Picture URL F(Optional)", 40),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    ws2 = wb.create_sheet(title='Sheet2')

    ws2.column_dimensions["A"].width = 30
    my_color = openpyxl.styles.colors.Color(rgb='00ffaa02')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    ws2['A1'].fill = my_fill
    ws2["A1"] = 'Product Category'
    ws2["A2"] = 'Golf'
    ws2["A3"] = 'Running'
    ws2["A4"] = 'Tennis'
    ws2["A5"] = 'Skiing & Snowboarding'
    ws2["A6"] = 'Biking'
    ws2["A7"] = 'Fishing'
    ws2["A8"] = 'Surfing/Kitesurfing'
    ws2["A9"] = 'Exploring'

    wb.save(response)
    return response

def import_view_broadmoor(request):

    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()
    return render(
        request,
        'vacay/upload_form_broadmoor.html',
        {
            'form': form,
            'title': 'Load Data',
            'header': 'Upload retail product info:'
        })


def import_broadmoor_data(request):

    global pro_num
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        if form.is_valid():

            if not request.session.exists(request.session.session_key):
                request.session.create()
            session_key = request.session.session_key
            session = Session.objects.get(session_key=session_key)
            idx = session.get_decoded().get('_auth_user_id')

            if idx is None:
                return redirect('/logout')

            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    pictureUrl = sheet.cell(r, 0).value
                    productName = sheet.cell(r, 1).value
                    inventory = sheet.cell(r, 2).value
                    category = sheet.cell(r, 3).value
                    additional = sheet.cell(r, 4).value

                    video = sheet.cell(r, 5).value
                    youtube = sheet.cell(r, 6).value
                    imageA = sheet.cell(r, 7).value
                    imageB = sheet.cell(r, 8).value
                    imageC = sheet.cell(r, 9).value
                    imageD = sheet.cell(r, 10).value
                    imageE = sheet.cell(r, 11).value
                    imageF = sheet.cell(r, 12).value

                    bproduct = BroadmoorProduct()
                    bproduct.adminID = idx
                    bproduct.bm_proName = productName
                    bproduct.bm_proInventoryNum = str(inventory).replace(".0","")
                    bproduct.bm_proCategory = category
                    bproduct.bm_proAdditional = additional

                    bproduct.bm_proImageUrl = pictureUrl

                    bproduct.video_url = video
                    bproduct.youtube_url = youtube
                    bproduct.imageA = imageA
                    bproduct.imageB = imageB
                    bproduct.imageC = imageC
                    bproduct.imageD = imageD
                    bproduct.imageE = imageE
                    bproduct.imageF = imageF

                    bproduct.save()
                    bproduct.bm_proid = bproduct.pk
                    bproduct.save()

                adminUser = AdminUser.objects.get(adminID=idx)
                bproducts = BroadmoorProduct.objects.filter(adminID=idx).order_by('-id')
                context = {'bproducts': bproducts, 'admin': adminUser}
                # cache.set('search_env', 3)
                return render(request, 'vacay/show_broadmoor_products.html', context)

            except XLRDError:
                return render(request, 'vacay/upload_form_broadmoor.html', {'note':'invalid_file'})
            except IOError:
                return render(request, 'vacay/upload_form_broadmoor.html', {'note': 'invalid_file'})
            except IndexError:
                return render(request, 'vacay/upload_form_broadmoor.html', {'note': 'invalid_file'})
            except DataError:
                return HttpResponse('Invalid file!')
        else:
            return render(request, 'vacay/upload_form_broadmoor.html', {'note':'invalid_file'})

    elif request.method == 'GET':
        pass

def export_xlsx_employee(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=employee_template.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Employees"

    row_num = 0

    columns = [
        (u"Picture Url", 40),
        (u"Name", 20),
        (u"Gender (from Sheet2)", 20),
        (u"Email", 30),
        (u"Password", 30),
        (u"Millennial/Gen Xer (from Sheet2)", 30),
        (u"VaCay Bucks Given", 30),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    ws2 = wb.create_sheet(title='Sheet2')

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 20
    my_color1 = openpyxl.styles.colors.Color(rgb='00ffaa02')
    my_fill1 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color1)
    my_color2 = openpyxl.styles.colors.Color(rgb='0000e5ff')
    my_fill2 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color2)

    ws2['A1'].fill = my_fill1
    ws2["A1"] = 'Gender'
    ws2["A2"] = 'Male'
    ws2["A3"] = 'Female'

    ws2['B1'].fill = my_fill2
    ws2["B1"] = 'Millennial/Gen Xer'
    ws2["B2"] = 'Millennial'
    ws2["B3"] = 'Gen Xer'

    wb.save(response)
    return response

def import_view_employee(request):

    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()
    return render(
        request,
        'vacay/upload_form_employee.html',
        {
            'form': form,
            'title': 'Load Data',
            'header': 'Upload employee information:'
        })

def import_employee_data(request):

    global pro_num
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        if form.is_valid():

            if not request.session.exists(request.session.session_key):
                request.session.create()
            session_key = request.session.session_key
            session = Session.objects.get(session_key=session_key)
            idx = session.get_decoded().get('_auth_user_id')

            if idx is None:
                return redirect('/logout')

            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    pictureUrl = sheet.cell(r, 0).value
                    name = sheet.cell(r, 1).value
                    gender = sheet.cell(r, 2).value
                    email = sheet.cell(r, 3).value
                    password = sheet.cell(r, 4).value
                    millennial = sheet.cell(r, 5).value
                    bucks = sheet.cell(r, 6).value

                    employee = Employee()
                    employee.adminID = idx
                    employee.em_name = name
                    employee.em_gender = gender
                    employee.em_email = email
                    employee.em_password = password
                    employee.em_millennial = millennial

                    if "$" in str(bucks):
                        employee.em_givenbuck = str(bucks)
                    else:
                        employee.em_givenbuck = "$" + str(bucks)

                    employee.em_image = pictureUrl

                    employee.save()
                    employee.em_id = employee.pk
                    employee.save()

                adminUser = AdminUser.objects.get(adminID=idx)
                employees = Employee.objects.filter(adminID=idx).order_by('-id')
                context = {'employees': employees, 'admin': adminUser}
                # cache.set('search_env', 3)
                return render(request, 'vacay/show_employees.html', context)

            except XLRDError:
                return render(request, 'vacay/upload_form_employee.html', {'note':'invalid_file'})
            except IOError:
                return render(request, 'vacay/upload_form_employee.html', {'note': 'invalid_file'})
            except IndexError:
                return render(request, 'vacay/upload_form_employee.html', {'note': 'invalid_file'})
            except DataError:
                return HttpResponse('Invalid file!')
        else:
            return render(request, 'vacay/upload_form_employee.html', {'note':'invalid_file'})

    elif request.method == 'GET':
        pass

def export_xlsx_job(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=job_template_xlsx.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Jobs"

    row_num = 0

    columns = [
        (u"Job Title", 30),
        (u"Req ID", 20),
        (u"Department", 30),
        (u"Location", 30),
        (u"Posting Date", 15),
        (u"Extra", 40),
        (u"Survey long link", 40),
        (u"Description", 40),
        (u"Video URL(Optional)", 40),
        (u"YouTube Video ID(Optional)", 40),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    wb.save(response)
    return response

def import_view_job(request):

    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()
    return render(
        request,
        'vacay/upload_form_job.html',
        {
            'form': form,
            'title': 'Load Data',
            'header': 'Upload job information:'
        })


def import_job_data(request):

    global pro_num
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        if form.is_valid():

            if not request.session.exists(request.session.session_key):
                request.session.create()
            session_key = request.session.session_key
            session = Session.objects.get(session_key=session_key)
            idx = session.get_decoded().get('_auth_user_id')

            if idx is None:
                return redirect('/logout')

            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    jobtitle = sheet.cell(r, 0).value
                    reqid = sheet.cell(r, 1).value
                    department = sheet.cell(r, 2).value
                    location = sheet.cell(r, 3).value
                    postingdate = sheet.cell(r, 4).value
                    extra = sheet.cell(r, 5).value
                    survey = sheet.cell(r, 6).value
                    description = sheet.cell(r, 7).value

                    video = sheet.cell(r, 8).value
                    youtube = sheet.cell(r, 9).value

                    date = strftime("%Y-%m-%d", gmtime())

                    job = Job()
                    job.adminID = idx
                    job.job_name = jobtitle
                    job.job_req = str(reqid).replace(".0","")
                    job.job_department = department
                    job.job_location = location
                    job.job_postdate = date
                    job.job_description = description
                    job.job_empty = extra
                    job.job_survey = survey

                    job.video_url = video
                    job.youtube_url = youtube

                    job.save()
                    job.job_id = job.pk
                    job.save()

                adminUser = AdminUser.objects.get(adminID=idx)
                jobs = Job.objects.filter(adminID=idx).order_by('-id')
                context = {'jobs': jobs, 'admin': adminUser}
                # cache.set('search_env', 3)
                return render(request, 'vacay/show_jobs.html', context)

            except XLRDError:
                return render(request, 'vacay/upload_form_job.html', {'note':'invalid_file'})
            except IOError:
                return render(request, 'vacay/upload_form_job.html', {'note': 'invalid_file'})
            except IndexError:
                return render(request, 'vacay/upload_form_job.html', {'note': 'invalid_file'})
            except DataError:
                return HttpResponse('Invalid file!')
        else:
            return render(request, 'vacay/upload_form_job.html', {'note':'invalid_file'})

    elif request.method == 'GET':
        pass

def export_xlsx_announce(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=announce_template_xlsx.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Announces"

    row_num = 0

    columns = [
        (u"Picture Url", 40),
        (u"Title", 30),
        (u"Audience", 30),
        (u"Subject", 30),
        (u"Call of Action", 30),
        (u"Message owner's Email", 30),
        (u"Survey long link", 40),
        (u"Description", 40),
        (u"Video URL(Optional)", 40),
        (u"YouTube Video ID(Optional)", 40),
        (u"Extra Picture URL A(Optional)", 40),
        (u"Extra Picture URL B(Optional)", 40),
        (u"Extra Picture URL C(Optional)", 40),
        (u"Extra Picture URL D(Optional)", 40),
        (u"Extra Picture URL E(Optional)", 40),
        (u"Extra Picture URL F(Optional)", 40),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    wb.save(response)
    return response

def import_view_announce(request):

    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()
    return render(
        request,
        'vacay/upload_form_announce.html',
        {
            'form': form,
            'title': 'Load Data',
            'header': 'Upload announcement information:'
        })

def import_announce_data(request):

    global pro_num
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        if form.is_valid():

            if not request.session.exists(request.session.session_key):
                request.session.create()
            session_key = request.session.session_key
            session = Session.objects.get(session_key=session_key)
            idx = session.get_decoded().get('_auth_user_id')

            if idx is None:
                return redirect('/logout')

            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    pictureUrl = sheet.cell(r, 0).value
                    title = sheet.cell(r, 1).value
                    audience = sheet.cell(r, 2).value
                    subject = sheet.cell(r, 3).value
                    callofaction = sheet.cell(r, 4).value
                    owneremail = sheet.cell(r, 5).value
                    survey = sheet.cell(r, 6).value
                    description = sheet.cell(r, 7).value

                    video = sheet.cell(r, 8).value
                    youtube = sheet.cell(r, 9).value
                    imageA = sheet.cell(r, 10).value
                    imageB = sheet.cell(r, 11).value
                    imageC = sheet.cell(r, 12).value
                    imageD = sheet.cell(r, 13).value
                    imageE = sheet.cell(r, 14).value
                    imageF = sheet.cell(r, 15).value

                    date = strftime("%Y-%m-%d", gmtime())

                    announce = Announce()
                    announce.adminID = idx
                    announce.an_title = title
                    announce.an_audience = audience
                    announce.an_subject = subject
                    announce.an_description = description
                    announce.an_callofaction = callofaction
                    announce.an_owneremail = owneremail
                    announce.an_survey = survey
                    announce.an_postdate = date

                    announce.an_image = pictureUrl

                    announce.video_url = video
                    announce.youtube_url = youtube
                    announce.imageA = imageA
                    announce.imageB = imageB
                    announce.imageC = imageC
                    announce.imageD = imageD
                    announce.imageE = imageE
                    announce.imageF = imageF

                    announce.save()
                    announce.an_id = announce.pk
                    announce.save()

                adminUser = AdminUser.objects.get(adminID=idx)
                announces = Announce.objects.filter(adminID=idx).order_by('-id')
                context = {'announces': announces, 'admin': adminUser}
                # cache.set('search_env', 3)
                return render(request, 'vacay/show_announcements.html', context)

            except XLRDError:
                return render(request, 'vacay/upload_form_announce.html', {'note':'invalid_file'})
            except IOError:
                return render(request, 'vacay/upload_form_announce.html', {'note': 'invalid_file'})
            except IndexError:
                return render(request, 'vacay/upload_form_announce.html', {'note': 'invalid_file'})
            except DataError:
                return HttpResponse('Invalid file!')
        else:
            return render(request, 'vacay/upload_form_announce.html', {'note':'invalid_file'})

    elif request.method == 'GET':
        pass


def edit_admin(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
            return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    context={'admin': adminUser}
    return render(request, 'vacay/edit_admin_profile.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_admin(request, admin_id):
    if request.method == 'POST':
        name = request.POST.get('name', None)
        email = request.POST.get('email', None)
        password = request.POST.get('password', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=admin_id)
        adminUser.adminName = name
        adminUser.adminEmail = email
        pwd = adminUser.adminPassword
        adminUser.adminPassword = password

        try:
            # image = request.FILES['photo']
            # fs = FileSystemStorage()
            # filename = fs.save(image.name, image)
            # image_url = fs.url(filename)
            # provider.proProfileImageUrl = settings.URL + image_url

            image = request.FILES['photo']
            fs = FileSystemStorage()
            try:
                x = request.POST.get('x', '0')
                y = request.POST.get('y', '0')
                w = request.POST.get('w', '32')
                h = request.POST.get('h', '32')
                #  return HttpResponse(w)
                file = profile_process(image, x, y, w, h)
                image = file
                # return HttpResponse('Cropped!')

            except MultiValueDictKeyError:
                print('No cropping')
            except ValueError:
                print('No cropping')

            filename = fs.save(image.name, image)
            uploaded_file_url = fs.url(filename)
            adminUser.adminImageUrl = settings.URL + uploaded_file_url

        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")


        logo = request.POST.get('b642', None)
        try:
            logo = request.FILES['logo']
            fs = FileSystemStorage()
            filename = fs.save(logo.name, logo)
            logo_url = fs.url(filename)
            adminUser.adminLogoImageUrl = settings.URL + logo_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            company = request.POST.get('company', None)
            if company is None:
                adminUser.adminCompany = adminUser.adminCompany
            else:
                adminUser.adminCompany = company

        except MultiValueDictKeyError or FileNotFoundError:
            print("Item Not Exist")
        except IntegrityError:
            adminUser.adminCompany = adminUser.adminCompany

        adminUser.save()

        user1 = authenticate(username=email, password=pwd)
        if user1 is not None:
            user1.password = password
            user1.set_password(password)
            user1.save()

        adminUser = AdminUser.objects.get(adminID=idx)

        if adminUser.adminBroadmoor == '0':
            all_providers = Provider.objects.filter(adminID=idx)
            context = {'all_providers': all_providers, 'admin': adminUser}
            return render(request, 'vacay/index.html', context)
        elif adminUser.adminBroadmoor == '1':
            bproducts = BroadmoorProduct.objects.filter(adminID=idx)
            context = {'bproducts': bproducts, 'admin': adminUser}
            return render(request, 'vacay/show_broadmoor_products.html', context)
        elif adminUser.adminBroadmoor == '2':
            employees = Employee.objects.filter(adminID=idx)
            context = {'employees': employees, 'admin': adminUser}
            # cache.set('search_env', 3)
            return render(request, 'vacay/show_employees.html', context)
        else:
            context = {'admin': adminUser}
            return render(request, 'vacay/edit_admin_profile.html', context)

    elif request.method == 'GET':
        pass

def goto_back(request):
    global serv
    try:
        serv = request.session['serv']
    except:
        print('Null')
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)

    if serv==1:
        return redirect('login_user_view')
    elif ret==1:
        return redirect('login_user_view')
    elif com==1:
        return redirect('login_user_view')

    if adminUser.adminBroadmoor == '0':
        all_providers = Provider.objects.filter(adminID=idx).order_by('-id')
        context = {'all_providers': all_providers, 'admin': adminUser}
        return render(request, 'vacay/index.html', context)
    elif adminUser.adminBroadmoor == '1':
        bproducts = BroadmoorProduct.objects.filter(adminID=idx).order_by('-id')
        context = {'bproducts': bproducts, 'admin': adminUser}
        return render(request, 'vacay/show_broadmoor_products.html', context)
    elif adminUser.adminBroadmoor == '2':
        employees = Employee.objects.filter(adminID=idx).order_by('-id')
        context = {'employees': employees, 'admin': adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_employees.html', context)
    else:
        context = {'admin': adminUser}
        return render(request, 'vacay/edit_admin_profile.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_service_provider_info(request):

    pickup_dict = {}
    pickup_records = []

    if request.method == 'POST':

        category = request.POST.get('proBeautyCategory', None)
        city = request.POST.get('proCity', None)

        providers = Provider.objects.filter(proCity=city)

        for provider in providers:

            provider = Provider.objects.get(proid=provider.proid)

            services = Service.objects.filter(proBeautyCategory=category, proid=provider.proid)
            for service in services:

                provider_id = provider.proid
                admin_id = provider.adminID
                provider_imageUrl = provider.proProfileImageUrl
                provider_firstName = provider.proFirstName
                provider_lastName = provider.proLastName
                provider_email = provider.proEmail
                provider_password = provider.proPassword
                provider_phone = provider.proPhone
                provider_city = provider.proCity
                provider_address = provider.proAddress
                provider_company = provider.proCompany
                provider_token = provider.proToken
                provider_servicePercent = provider.proServicePercent
                provider_salary = provider.proSalary
                provider_productSalePercent = provider.proProductSalePercent
                provider_proAvailable = provider.proAvailable

                service = Service.objects.get(serviceid=service.serviceid)

                service_id = service.serviceid
                service_imageUrl = service.proServicePictureUrl
                service_beautyCategory = service.proBeautyCategory
                service_beautySubCategory = service.proBeautySubCategory
                service_price = service.proServicePrice
                service_description = service.proServiceDescription

                record = {
                    "serviceid": service_id,
                    "proid": provider_id,
                    "adminID": admin_id,
                    "proServicePictureUrl": service_imageUrl,
                    "proBeautyCategory": service_beautyCategory,
                    "proBeautySubcategory": service_beautySubCategory,
                    "proServicePrice": service_price,
                    "proServiceDescription": service_description,
                    "proProfileImageUrl": provider_imageUrl,
                    "proFirstName": provider_firstName,
                    "proLastName": provider_lastName,
                    "proEmail": provider_email,
                    "proPassword": provider_password,
                    "proPhone": provider_phone,
                    "proCity": provider_city,
                    "proAddress": provider_address,
                    "proCompany": provider_company,
                    "proToken": provider_token,
                    "proServicePercent": provider_servicePercent,
                    "proSalary": provider_salary,
                    "proProductSalePercent": provider_productSalePercent,
                    "proAvailable": provider_proAvailable
                }

                pickup_records.append(record)

        pickup_dict["service_provider_info"] = pickup_records

        resp = {'result_code': '0', "service_provider_info": pickup_records}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':

        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_provider_schedule(request):

    if request.method == 'POST':

        proid = request.POST.get('proid', None)

        schedules = ProviderSchedule.objects.filter(proid=proid)

        serializer = ProviderScheduleSerializer(schedules, many=True)
        resp = {'result_code': '0', 'available_info': serializer.data, }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_product_info(request):

    if request.method == 'POST':

        proid = request.POST.get('proid', None)

        products = Product.objects.filter(proid=proid)

        serializer = ProductSerializer(products, many=True)
        resp = {'result_code': '0', 'productInfo': serializer.data, }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_providers_by_adminID(request):

    if request.method == 'POST':

        admin_id = request.POST.get('adminID', None)

        providers = Provider.objects.filter(adminID=admin_id)

        serializer = ProviderSerializer(providers, many=True)
        resp = {'result_code': '0', 'provider_info': serializer.data, }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def increase_interaction(request):

    if request.method == 'POST':

        employeeid = request.POST.get('em_id', None)

        employee = Employee.objects.get(em_id=employeeid)
        if employee.em_interaction == '':
            employee.em_interaction = '1'
        elif employee.em_interaction == '0':
            employee.em_interaction = '1'
        else:
            employee.em_interaction = int(employee.em_interaction) + 1
        employee.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_provider_token(request):

    if request.method == 'POST':

        proid = request.POST.get('proid', None)
        token = request.POST.get('proToken', None)

        provider = Provider.objects.get(proid=proid)
        provider.proToken = token
        provider.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_broadmoor_info(request):

    if request.method == 'POST':

        pickup_dict = {}
        pickup_records = []

        bcategory = request.POST.get('bm_proCategory', None)

        bproducts = BroadmoorProduct.objects.filter(bm_proCategory=bcategory)

        for bproduct in bproducts:

            bproid = bproduct.bm_proid
            adminid = bproduct.adminID
            adminemail = AdminUser.objects.get(adminID=adminid).adminEmail
            adminlogo = AdminUser.objects.get(adminID=adminid).adminLogoImageUrl
            bproimage = bproduct.bm_proImageUrl
            bproname = bproduct.bm_proName
            bproinventorynum = bproduct.bm_proInventoryNum
            bprocategory = bproduct.bm_proCategory
            bproadditional = bproduct.bm_proAdditional

            record = {
                "bm_proid": bproid,
                "adminID": adminid,
                "adminEmail": adminemail,
                "adminLogoImageUrl": adminlogo,
                "bm_proImageUrl": bproimage,
                "bm_proName": bproname,
                "bm_proInventoryNum": bproinventorynum,
                "bm_proCategory": bprocategory,
                "bm_proAdditional": bproadditional,
            }

            pickup_records.append(record)

        pickup_dict["broadmoor_info"] = pickup_records

        resp = {'result_code': '0', "broadmoor_info": pickup_records}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_broadmoor_detail(request):

    if request.method == 'POST':

        bproid = request.POST.get('bm_proid', None)

        bprodetails = BroadmoorProductDetail.objects.filter(bm_proid=bproid)

        serializer = BroadmoorProductDetailSerializer(bprodetails, many=True)
        resp = {'result_code': '0', 'detail_info': serializer.data, }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_employees_by_adminID(request):

    if request.method == 'POST':

        pickup_dict = {}
        pickup_records = []

        admin_id = request.POST.get('adminID', None)

        employees = Employee.objects.filter(adminID=admin_id)

        for employee in employees:

            emid = employee.em_id
            adminid = employee.adminID
            admincompany = AdminUser.objects.get(adminID=adminid).adminCompany
            adminlogo = AdminUser.objects.get(adminID=adminid).adminLogoImageUrl
            emimage = employee.em_image
            emname = employee.em_name
            emgender = employee.em_gender
            ememail = employee.em_email
            empassword = employee.em_password
            emmillennial = employee.em_millennial
            emgivenbuck = employee.em_givenbuck
            emusedbuck = employee.em_usedbuck
            eminteraction = employee.em_interaction
            emstatus = employee.em_status

            record = {
                "em_id": emid,
                "adminID": adminid,
                "adminCompany": admincompany,
                "adminLogoImageUrl": adminlogo,
                "em_image": emimage,
                "em_name": emname,
                "em_gender": emgender,
                "em_email": ememail,
                "em_password": empassword,
                "em_millennial": emmillennial,
                "em_givenbuck": emgivenbuck,
                "em_usedbuck": emusedbuck,
                "em_interaction": eminteraction,
                "em_status": emstatus,
            }

            pickup_records.append(record)

        pickup_dict["employee_info"] = pickup_records

        resp = {'result_code': '0', "employee_info": pickup_records}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_employee_status(request):

    if request.method == 'POST':

        emid = request.POST.get('em_id', None)

        employee = Employee.objects.get(em_id=emid)
        employee.em_status = '1'
        employee.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_companies(request):

    if request.method == 'POST':

        pickup_dict = {}
        pickup_records = []

        adminUsers = AdminUser.objects.all().order_by('adminCompany')
        company_init = adminUsers[:1].get().adminCompany

        for adminUser in adminUsers:
            if adminUser.adminCompany != '' and adminUser.adminCompany != company_init:
                company_init=adminUser.adminCompany
                record = {
                    "adminID": adminUser.adminID,
                    "adminCompany": adminUser.adminCompany,
                    "adminLogoImageUrl": adminUser.adminLogoImageUrl,
                }

                pickup_records.append(record)


        resp = {'result_code': '0', 'company_info': pickup_records, }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_jobs(request):

    if request.method == 'POST':

        pickup_dict = {}
        pickup_records = []

        admin_id = request.POST.get('adminID', None)

        jobs = Job.objects.filter(adminID=admin_id)

        for job in jobs:

            jobid = job.job_id
            adminid = job.adminID
            admincompany = AdminUser.objects.get(adminID=adminid).adminCompany
            adminlogo = AdminUser.objects.get(adminID=adminid).adminLogoImageUrl
            name = job.job_name
            req = job.job_req
            department = job.job_department
            location = job.job_location
            description = job.job_description
            postdate = job.job_postdate
            empty = job.job_empty
            survey = job.job_survey

            record = {
                "job_id": jobid,
                "adminID": adminid,
                "adminCompany": admincompany,
                "adminLogoImageUrl": adminlogo,
                "job_name": name,
                "job_req": req,
                "job_department": department,
                "job_location": location,
                "job_description": description,
                "job_postdate": postdate,
                "job_empty": empty,
                "job_survey":survey,
            }

            pickup_records.append(record)

        pickup_dict["job_info"] = pickup_records

        resp = {'result_code': '0', "job_info": pickup_records}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_announces(request):

    if request.method == 'POST':

        pickup_dict = {}
        pickup_records = []

        admin_id = request.POST.get('adminID', None)

        announces = Announce.objects.filter(adminID=admin_id)

        for announce in announces:

            anid = announce.an_id
            adminid = announce.adminID
            admincompany = AdminUser.objects.get(adminID=adminid).adminCompany
            adminlogo = AdminUser.objects.get(adminID=adminid).adminLogoImageUrl
            image = announce.an_image
            title = announce.an_title
            audience = announce.an_audience
            subject = announce.an_subject
            description = announce.an_description
            callofaction = announce.an_callofaction
            owneremail = announce.an_owneremail
            viewnum = announce.an_viewnum
            responsenum = announce.an_responsenum
            postdate = announce.an_postdate
            survey = announce.an_survey

            record = {
                "an_id": anid,
                "adminID": adminid,
                "adminCompany": admincompany,
                "adminLogoImageUrl": adminlogo,
                "an_image": image,
                "an_title": title,
                "an_audience": audience,
                "an_subject": subject,
                "an_description": description,
                "an_postdate": postdate,
                "an_callofaction": callofaction,
                "an_owneremail": owneremail,
                "an_viewnum": viewnum,
                "an_responsenum": responsenum,
                "an_survey": survey,
            }

            pickup_records.append(record)

        pickup_dict["announce_info"] = pickup_records

        resp = {'result_code': '0', "announce_info": pickup_records}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_announce_view(request):

    if request.method == 'POST':

        emid = request.POST.get('em_id', None)
        anid = request.POST.get('an_id', None)
        indexx = request.POST.get('index', None)

        announce = Announce.objects.get(an_id=anid)
        if indexx =='0':
            if announce.an_viewnum == '':
                announce.an_viewnum = '1'
            elif announce.an_viewnum == '0':
                announce.an_viewnum = '1'
            else:
                announce.an_viewnum = int(announce.an_viewnum) + 1
        else:
            if announce.an_responsenum == '':
                announce.an_responsenum = '1'
            elif announce.an_responsenum == '0':
                announce.an_responsenum = '1'
            else:
                announce.an_responsenum = int(announce.an_responsenum) + 1

        announce.save()

        anview = AnnounceView()
        anview.em_id = emid
        anview.an_id = anid

        if indexx =='1':
            anview.is_signup = 'yes'

        anview.save()
        anview.v_id=anview.pk
        anview.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_employees_for_announce(request):

    if request.method == 'POST':

        pickup_dict = {}
        pickup_records = []

        admin_id = request.POST.get('adminID', None)
        anid = request.POST.get('an_id', None)

        employees = Employee.objects.filter(adminID=admin_id)

        for employee in employees:

            anviews = AnnounceView.objects.filter(em_id=employee.em_id, an_id=anid)
            if anviews.count()>0:
                emid = employee.em_id
                adminid = employee.adminID
                admincompany = AdminUser.objects.get(adminID=adminid).adminCompany
                adminlogo = AdminUser.objects.get(adminID=adminid).adminLogoImageUrl
                emimage = employee.em_image
                name = employee.em_name
                email = employee.em_email

                record = {
                    "em_id": emid,
                    "em_image": emimage,
                    "em_name": name,
                    "em_email": email
                }

                pickup_records.append(record)

        pickup_dict["em_info"] = pickup_records

        resp = {'result_code': '0', "em_info": pickup_records}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login_provider(request):

    if request.method == 'POST':

        email = request.POST.get('proEmail', None)

        providers = Provider.objects.filter(proEmail=email)
        count = providers.count()
        if count>0:

            serializer = ProviderSerializer(providers, many=True)
            resp = {'result_code': '0', 'provider_info': serializer.data}

            return JsonResponse(resp, status=status.HTTP_200_OK)

            # user = authenticate(username=email, password=password)
            # if user is not None:
            #     login(request, user)
            #     if not request.session.exists(request.session.session_key):
            #         request.session.create()
            #     session_key = request.session.session_key
            #     session = Session.objects.get(session_key=session_key)
            #     idx = session.get_decoded().get('_auth_user_id')
            #     user = User.objects.get(pk=idx)
            #     eml = user.email
            #
            #     serializer = AdminUserSerializer(user0, many=True)
            #     resp = {'result_code': '0', 'adminData': serializer.data}
            #
            #     return JsonResponse(resp, status=status.HTTP_200_OK)
            #
            # else:
            #     resp = {'result_code': '113'}
            #     return JsonResponse(resp, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':

        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_services_from_provider(request):

    if request.method == 'POST':

        proid = request.POST.get('proid', None)
        provider = Provider.objects.get(proid=proid)

        processing_fee = 0.2
        servicePercent = float(str(provider.proServicePercent).replace("%", "").replace(",", ""))
        managerPercent = 100 - servicePercent

        services = Service.objects.filter(proid=proid)

        for service in services:
            totalPrice = float(str(service.proServicePrice).replace("$", "").replace(",", "")) - (float(str(service.proServicePrice).replace("$", "").replace(",", "")) * processing_fee)
            providerTakeHome = servicePercent * totalPrice * 0.01
            managerTakeHome = totalPrice - providerTakeHome

            service.providerTakeHome = str(providerTakeHome)[0:8]
            service.managerTakeHome = str(managerTakeHome)[0:8]

            service.save()

        services = Service.objects.filter(proid=proid)

        serializer = ServiceSerializer(services, many=True)
        resp = {'result_code': '0', 'service_info': serializer.data}

        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':

        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_products_from_provider(request):

    if request.method == 'POST':

        proid = request.POST.get('proid', None)
        provider = Provider.objects.get(proid=proid)

        processing_fee = 0.2
        providerPercent = float(str(provider.proProductSalePercent).replace("%", "").replace(",", ""))
        managerPercent = 100 - providerPercent

        products = Product.objects.filter(proid=proid)

        for product in products:
            totalPrice = float(str(product.itemPrice).replace("$", "").replace(",", "")) - (float(str(product.itemPrice).replace("$", "").replace(",", "")) * processing_fee)
            providerTakeHome = providerPercent * totalPrice * 0.01
            managerTakeHome = totalPrice - providerTakeHome

            product.providerTakeHome = str(providerTakeHome)[0:8]
            product.managerTakeHome = str(managerTakeHome)[0:8]

            product.save()

        products = Product.objects.filter(proid=proid)

        serializer = ProductSerializer(products, many=True)
        resp = {'result_code': '0', 'productInfo': serializer.data}

        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':

        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_provider_schedule(request):

    if request.method == 'POST':

        proid = request.POST.get('proid', None)
        start = request.POST.get('availableStart', None)
        end = request.POST.get('availableEnd', None)
        comment = request.POST.get('availableComment', None)

        schedule = ProviderSchedule()
        schedule.proid = proid
        schedule.availableStart = start
        schedule.availableEnd = end
        schedule.availableComment = comment
        schedule.save()
        schedule.availableid = schedule.pk
        schedule.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_provider_schedule(request):

    if request.method == 'POST':

        availableid = request.POST.get('availableid', None)

        ProviderSchedule.objects.get(availableid=availableid).delete()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login_employee(request):

    if request.method == 'POST':

        pickup_records =[]
        email = request.POST.get('em_email', None)

        employees = Employee.objects.filter(em_email=email)
        count = employees.count()
        if count>0:

            for employee in employees:
                emid = employee.em_id
                adminid = employee.adminID
                admincompany = AdminUser.objects.get(adminID=adminid).adminCompany
                adminlogo = AdminUser.objects.get(adminID=adminid).adminLogoImageUrl
                image = employee.em_image
                name = employee.em_name
                gender = employee.em_gender
                email = employee.em_email
                password = employee.em_password
                millennial = employee.em_millennial
                givenbuck = employee.em_givenbuck
                usedbuck = employee.em_usedbuck
                interaction = employee.em_interaction
                emstatus = employee.em_status

                record = {
                    "em_id": emid,
                    "adminID": adminid,
                    "adminCompany": admincompany,
                    "adminLogoImageUrl": adminlogo,
                    "em_image": image,
                    "em_name": name,
                    "em_gender": gender,
                    "em_email": email,
                    "em_password": password,
                    "em_millennial": millennial,
                    "em_givenbuck": givenbuck,
                    "em_usedbuck": usedbuck,
                    "em_interaction": interaction,
                    "em_status": emstatus,
                }

                pickup_records.append(record)

            resp = {'result_code': '0', "employee_info": pickup_records}
            return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':

        pass

def sendSimpleEmail(request,employee_id):
    employee = Employee.objects.get(em_id=employee_id)
    fromaddress = 'cayley@vacaycarpediem.com'             #     cayley@vacaycarpediem.com
    toaddress =  employee.em_email
    subject = 'Welcome to VaCay!'
    body = ''

    if "$" in employee.em_givenbuck:
        emgivenbuck = employee.em_givenbuck
    else:
        emgivenbuck = "$"+employee.em_givenbuck

    html = """\
        <html>
          <head></head>
          <body>
          <center>
            <p style="font-weight:600;font-size:25px; color:#000000; font-family:verdana;">VaCay Welcomes You!</p>
            <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/vacaylogo.jpg" style="width:100px;height:100px;border-radius: 20%;"/></a>
            </center>
            <h2 style="margin-left:20px; color: #e69900;">Your company has signed you up!</h2>
            <div style="font-size:16px; font-family:verdana; color:black;">
                VaCay's mission is to help you feel refreshed, rejuvenated and know how appreciated you are after each use!<br><br>
                Please get to know your <label style="font-weight:600;">colleagues, make new friends</label>, check out the recommended <label style="font-weight:600;">top restaurants</label>, enjoy your on-demand
                <label style="font-weight:600;">Employee Incentives</label>, have a wonderful time finding your colleagues to <label style="font-weight:600;">play activities and explore with</label>
                 and learn about your company's <label style="font-weight:600;">Jobs</label> and
                <label style="font-weight:600;">Announcements</label>!<br><br>
                Your company has given you {givenbuck} <label style="font-weight:600; color:#e69900;">VaCay Bucks</label> <img src="https://www.vacayadmin.com/static/vacay/images/bucklogo.png" style="width:150px; height:70px;"><br><br>
                which can be used for the <label style="font-weight:600;">On-Demand Employee Appreciation Services</label>. <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/beautyservicebackground.jpg" style="width:90px; height:90px;"></a><br><br>
                Your username is <a href="mailto:{username}">{username}</a> and we will send you another email with your password and how to login.<br><br>
                Have a wonderful day!<br><br>
                Warmly,<br><br>
                The VaCay Team<br><br>
                <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/logo.jpg" style="width:140px; height:70px;"></a>
            </div>
          </body>
        </html>
        """
    if emgivenbuck.endswith('.0') or emgivenbuck.endswith('.00'):
        emgivenbuck = emgivenbuck.replace('.0', '').replace('.00', '')
    html = html.format(givenbuck=emgivenbuck, username=employee.em_email)

    html2 = """\
            <html>
              <head></head>
              <body>
              <center>
                <p style="font-weight:600;font-size:25px; color:#000000; font-family:verdana;">VaCay Welcomes You!</p>
                <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/vacaylogo.jpg" style="width:100px;height:100px;border-radius: 20%;"/></a>
                </center>
                <h2 style="margin-left:20px; color: #e69900;">Your company has signed you up!</h2>
                <div style="font-size:16px; color:black;">
                    Please click <a href="https://www.vacayalldays.com/employee_login_page?csrfmiddlewaretoken=38zU9eycA2RneFE4UqoYpggV6NLwWq7g2RgluAvKMG6JxvCfuE9hens0TaFbun4U"><label style="color:#e69900; font-weight:600;">here</label></a> to login.<br><br>
                    If you logout, you'll need to click "Employee" <a href="https://www.vacayalldays.com"><button style="width:150px; height:40px; text-align:center; font-size:16px; font-weight:600; color:white; background:green; border-radius:50px;">Employee</button></a> to login again.<br><br>
                    You'll enter in your username <a href="mailto:{username}"><label style="font-weight:500;">{username}</label></a> and the password <label style="font-weight:600;">{password}</label>.<br><br>
                    You'll also be able to review any FAQs here <a href="https://www.vacayalldays.com"><button style="width:150px; height:40px; text-align:center; font-size:16px; font-weight:600; color:white; background:green; border-radius:50px;">FAQs</button></a> and any Jobs, here <a href="https://www.vacayalldays.com"><button style="width:150px; height:40px; text-align:center; font-size:16px; font-weight:600; color:white; background:green; border-radius:50px;">Jobs</button></a>.<br><br>
                    Thank you!<br><br>
                    Warmly,<br>
                    The VaCay Team<br><br>
                    <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/logo.jpg" style="width:140px; height:70px;"></a>
                </div>
              </body>
            </html>
            """
    html2 = html2.format(password=employee.em_password, username=employee.em_email)

    html1 = """\
    <html>
      <head></head>
      <body>
        <p>Please visit to <a href="http://www.vacaycarpediem.com" style="font-size:17px; font-style:italic;">VaCay</a>
        </p>
      </body>
    </html>
    """

    text = "VaCay's mission is to help you feel refreshed, rejuvenated and know how appreciated you are after each use!\nPlease get to know your colleagues, " \
           "make new friends, check out the recommended top restaurants, enjoy your on-demand Employee Incentives, have a wonderful time finding your colleagues to play activities and " \
           "explore with and learn about your company's Jobs and Announcements!\n" \
           "Your company has given you " + emgivenbuck + " VaCay bucks which can be used for the On-Demand Employee Appreciation Services.\n" \
           "Your username is " + "manish@vacaycarpediem.com" +" and we will send you another email with your password and how to login.\n" \
                                                                "Have a wonderful day!\nWarmly,\nThe VaCay Team\n"

    text2 = "Here is your password: " + employee.em_password + "\nWarmly,\nThe VaCay Team\nBelow are some instructions to help you best use VaCay. Please click on the link to learn more. \n"

    username = settings.EMAIL_HOST_USER
    password = settings.EMAIL_HOST_PASSWORD
    msg = MIMEMultipart()
    msg['From']=fromaddress
    msg['To'] = toaddress
    msg['Subject'] = subject
    # msg['Body'] = body
    # message = """From: %s\nTo: %s\nSubject: %s\n\n%s
    # """ % (fromaddress, ", ".join(toaddress), subject, body)

    # body = MIMEText(html, 'html')
    body1 = MIMEText(html, 'html')
    body2 = MIMEText(html1, 'html')
    body4 = MIMEText(html2, 'html')
    body = MIMEText(text, 'plain')
    body3 = MIMEText(text2, 'plain')

    msg.attach(body1)
    # msg.attach(body)
    # msg.attach(body2)

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    employees = Employee.objects.filter(adminID=idx).order_by('-id')

    try:
        server = smtplib.SMTP('mail.smtp2go.com', 2525)
        server.ehlo()
        server.starttls()
        server.login(username, password)
        server.sendmail(fromaddress, toaddress, msg.as_string())

        msg2 = MIMEMultipart()
        msg2['From'] = fromaddress
        msg2['To'] = toaddress
        msg2['Subject'] = subject

        # msg2.attach(body1)
        # msg2.attach(body3)
        msg2.attach(body4)

        server.sendmail(fromaddress, toaddress, msg2.as_string())
        server.quit()

        employee.em_status = '1'
        employee.save()

        context = {'employees': employees, 'admin': adminUser, 'success': 'success'}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_employees.html', context)

        # return HttpResponse(toaddress + ': Sent!')
    except:

        context = {'employees': employees, 'admin': adminUser, 'failure': 'failure'}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_employees.html', context)

        # return HttpResponse("failed to send mail")


def show_on_map(request, provider_id):
    provider = Provider.objects.get(proid = provider_id)
    context = {'provider':provider}
    return render(request, 'vacay/search_address_onmap.html', context)

def show_my_loc(request):
    return render(request, 'vacay/show_on_map.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def nearby_services(request):

    if request.method == 'POST':

        lat = request.POST.get('latitude', None)
        lng = request.POST.get('longitude', None)
        service_type = request.POST.get('types', None)
        range = request.POST.get('ranges', None)
        address = request.POST.get('address', None)

        if service_type == 'All...':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address}
            return render(request, 'vacay/nearby_service.html', context)
        elif service_type == 'Hospital':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address}
            return render(request, 'vacay/nearby_hospital.html', context)
        elif service_type == 'Airport':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address}
            return render(request, 'vacay/nearby_airport.html', context)
        elif service_type == 'Restaurant':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address}
            return render(request, 'vacay/nearby_restaurant.html', context)
        elif service_type == 'Bank':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address}
            return render(request, 'vacay/nearby_bank.html', context)
        elif service_type == 'Beauty-Salon':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address}
            return render(request, 'vacay/nearby_beauty.html', context)
        elif service_type == 'Accounting':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'accounting'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Bar':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'bar'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Cafe':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'cafe'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Amusementpark':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'amusementpark'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Bookstore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'bookstore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Busstation':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'busstation'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Bicyclestore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'bicyclestore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Campground':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'campground'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Carrepair':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'carrepair'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Carrental':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'carrental'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Carwash':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'carwash'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Cardealer':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'cardealer'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Casino':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'casino'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Church':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'church'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Cityhall':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'cityhall'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Clothingstore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'clothingstore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Conveniencestore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'conveniencestore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Courthouse':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'courthouse'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Departmentstore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'departmentstore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Dentist':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'dentist'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Doctor':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'doctor'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Electrician':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'electrician'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Electronicsstore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'electronicsstore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Embassy':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'embassy'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Firestation':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'firestation'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Florist':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'florist'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'Furniturestore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'furniturestore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'gasstation':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'gasstation'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'gym':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'gym'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'haircare':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'haircare'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'hardwarestore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'hardwarestore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'hindutemple':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'hindutemple'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'homegoodsstore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'homegoodsstore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'aquarium':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'aquarium'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'artgallery':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'artgallery'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'atm':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'atm'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'bakery':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'bakery'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'bowlingalley':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'bowlingalley'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'insuranceagency':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'insuranceagency'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'jewelrystore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'jewelrystore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'laundry':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'laundry'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'lawyer':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'lawyer'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'library':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'library'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'liquorstore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'liquorstore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'localgovernmentoffice':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'localgovernmentoffice'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'locksmith':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'locksmith'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'lodging':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'lodging'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'mealdelivery':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'mealdelivery'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'mealtakeaway':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'mealtakeaway'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'mosque':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'mosque'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'movierental':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'movierental'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'movietheater':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'movietheater'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'movingcompany':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'movingcompany'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'museum':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'museum'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'nightclub':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'nightclub'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'painter':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'painter'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'park':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'park'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'parking':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'parking'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'petstore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'petstore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'pharmacy':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'pharmacy'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'physiotherapist':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'physiotherapist'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'placeofworship':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'placeofworship'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'plumber':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'plumber'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'police':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'police'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'postoffice':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'postoffice'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'realestateagency':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'realestateagency'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'roofingcontractor':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'roofingcontractor'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'rvpark':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'rvpark'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'shoestore':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'shoestore'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'shoppingmall':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'shoppingmall'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'spa':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'spa'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'stadium':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'stadium'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'storage':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'storage'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'store':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'store'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'subwaystation':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'subwaystation'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'synagogue':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'synagogue'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'taxistand':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'taxistand'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'trainstation':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'trainstation'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'transitstation':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'transitstation'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'travelagency':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'travelagency'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'university':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'university'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'veterinarycare':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'veterinarycare'}
            return render(request, 'vacay/nearby.html', context)
        elif service_type == 'zoo':
            context = {'latitude': str(lat), 'longitude': str(lng), 'range': str(range), 'address':address, 'type': 'zoo'}
            return render(request, 'vacay/nearby.html', context)

    elif request.method == 'GET':
        pass

def get_my_location(request):
    return render(request, 'vacay/myloc_onmap.html')

def get_all_providerSchedules(request):

    global serv2
    global prod

    serv2=request.session['serv2']
    prod=request.session['prod']

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser=AdminUser.objects.get(adminID=idx)
    all_providers=Provider.objects.filter(adminID=idx)

    for provider in all_providers:
        proSchedules = ProviderSchedule.objects.filter(proid = provider.proid)
        curDateTime = datetime.datetime.now()
        scheduleCount = proSchedules.count()
        if scheduleCount == 0:
            provider.proAvailable = 'true'
            provider.save()
            continue
        else:
            if scheduleCount > 7:
                last = 7
            else: last = scheduleCount
        for proSchedule in proSchedules[(scheduleCount-last):]:
            availableEnd = proSchedule.availableEnd
            availableStart = proSchedule.availableStart
            availableEndDateTime = datetime.datetime.strptime(availableEnd, "%B %d,%Y - %I:%M %p")
            availableStartDateTime = datetime.datetime.strptime(availableStart, "%B %d,%Y - %I:%M %p")
            if curDateTime < availableEndDateTime:
                if curDateTime > availableStartDateTime:
                    provider.proAvailable = 'false'
                    provider.save()
                    break
                else:
                    provider.proAvailable = 'true'
                    provider.save()
            else:
                provider.proAvailable = 'true'
                provider.save()

    all_providers = Provider.objects.filter(adminID=idx)

    if serv2 == 1:
        context = {'all_providers': all_providers, 'admin': adminUser, 'note': 'service'}
        return render(request, 'vacay/select_provider.html', context)
    elif prod == 1:
        context = {'all_providers': all_providers, 'admin': adminUser, 'note': 'product'}
        return render(request, 'vacay/select_provider.html', context)
    else:
        context = {'all_providers': all_providers, 'admin': adminUser}
        return render(request, 'vacay/index.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_provider_multiple_schedule(request):

    if request.method == 'POST':

        proid = request.POST.get('proid', None)
        schedulestr = request.POST.get('schedulestr', None)

        try:
            decoded = json.loads(schedulestr)
            for schedule_data in decoded['schedule']:

                start = schedule_data['start']
                end = schedule_data['end']
                comment = schedule_data['comment']

                schedule = ProviderSchedule()
                schedule.proid = proid
                schedule.availableStart = start
                schedule.availableEnd = end
                schedule.availableComment = comment
                schedule.save()
                schedule.availableid = schedule.pk
                schedule.save()

            resp = {'result_code': '0'}
            return JsonResponse(resp, status=status.HTTP_200_OK)

        except:
            resp = {'result_code': '1'}
            return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_all_providerSchedules(request):
    if request.method == 'POST':

        all_providers = Provider.objects.all()

        for provider in all_providers:
            proSchedules = ProviderSchedule.objects.filter(proid=provider.proid)
            curDateTime = datetime.datetime.now()
            scheduleCount = proSchedules.count()
            if scheduleCount == 0:
                provider.proAvailable = 'true'
                provider.save()
                continue
            else:
                if scheduleCount > 7:
                    last = 7
                else:
                    last = scheduleCount
            for proSchedule in proSchedules[(scheduleCount - last):]:
                availableEnd = proSchedule.availableEnd
                availableStart = proSchedule.availableStart
                availableEndDateTime = datetime.datetime.strptime(availableEnd, "%B %d,%Y - %I:%M %p")
                availableStartDateTime = datetime.datetime.strptime(availableStart, "%B %d,%Y - %I:%M %p")
                if curDateTime < availableEndDateTime:
                    if curDateTime > availableStartDateTime:
                        provider.proAvailable = 'false'
                        provider.save()
                        break
                    else:
                        provider.proAvailable = 'true'
                        provider.save()
                else:
                    provider.proAvailable = 'true'
                    provider.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)
    elif request.method == 'GET':
        pass

def select_provider_service(request):
    global serv2
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    serv2=1
    request.session['serv2']=1

    adminUser=AdminUser.objects.get(adminID=idx)
    all_providers=Provider.objects.filter(adminID=idx)

    context = {'all_providers': all_providers, 'admin': adminUser, 'note':'service'}
    # cache.set('search_env', 3)
    return render(request, 'vacay/select_provider.html', context)

def select_provider_product(request):
    global prod
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    prod=1
    request.session['prod']=1

    adminUser=AdminUser.objects.get(adminID=idx)
    all_providers=Provider.objects.filter(adminID=idx)

    context = {'all_providers': all_providers, 'admin': adminUser, 'note':'product'}
    # cache.set('search_env', 3)
    return render(request, 'vacay/select_provider.html', context)

def show_tips_tricks(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    tipstricks=TipsTricks.objects.filter(adminID=idx).order_by('-id')
    context = {'tipstricks': tipstricks, 'admin': adminUser}
    return render(request, 'vacay/show_tipstricks.html', context)

def add_tips_tricks(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    return render(request, 'vacay/add_tips_tricks.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_tiptrick_process(request):
    if request.method == 'POST':

        title = request.POST.get('title', None)
        audience = request.POST.get('audience', None)
        subject = request.POST.get('subject', None)
        callofaction = request.POST.get('callofaction', None)
        owneremail = request.POST.get('owneremail', None)
        description = request.POST.get('description', None)
        survey = request.POST.get('survey', None)
        youtubeurl = request.POST.get('youtubeurl', None)
        admin_free = request.POST.get('afree', None)
        buy_paid = request.POST.get('bpaid', None)
        free = request.POST.get('fr', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        date=strftime("%Y-%m-%d", gmtime())

        tipstricks = TipsTricks()
        tipstricks.adminID = idx
        tipstricks.title = title
        tipstricks.audience = audience
        tipstricks.subject = subject
        tipstricks.description = description
        tipstricks.callofaction = callofaction
        tipstricks.owneremail = owneremail
        tipstricks.postdate = date
        tipstricks.survey = survey
        tipstricks.op_admin = admin_free
        tipstricks.op_buy = buy_paid
        tipstricks.op_free = free

        try:
            image = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            tipstricks.image = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            tipstricks.video = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            tipstricks.youtube = youtubeurl

        photoa = request.POST.get('b64a', None)
        photob = request.POST.get('b64b', None)
        photoc = request.POST.get('b64c', None)
        photod = request.POST.get('b64d', None)
        photoe = request.POST.get('b64e', None)
        photof = request.POST.get('b64f', None)

        if photoa is not None:
            tipstricks.imageA = photoa
        if photob is not None:
            tipstricks.imageB = photob
        if photoc is not None:
            tipstricks.imageC = photoc
        if photod is not None:
            tipstricks.imageD = photod
        if photoe is not None:
            tipstricks.imageE = photoe
        if photof is not None:
            tipstricks.imageF = photof

        desca = request.POST.get('desca', None)
        descb = request.POST.get('descb', None)
        descc = request.POST.get('descc', None)
        descd = request.POST.get('descd', None)
        desce = request.POST.get('desce', None)
        descf = request.POST.get('descf', None)

        if desca is not None and desca != 'None':
            tipstricks.descA = desca
        if descb is not None and descb != 'None':
            tipstricks.descB = descb
        if descc is not None and descc != 'None':
            tipstricks.descC = descc
        if descd is not None and descd != 'None':
            tipstricks.descD = descd
        if desce is not None and desce != 'None':
            tipstricks.descE = desce
        if descf is not None and descf != 'None':
            tipstricks.descF = descf

        tipstricks.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        tipstricks = TipsTricks.objects.filter(adminID=idx).order_by('-id')
        context = {'tipstricks': tipstricks, 'admin':adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_tipstricks.html', context)

    elif request.method == 'GET':
        pass

def delete_tiptrick(request, tiptrick_id):
    TipsTricks.objects.get(id=tiptrick_id).delete()

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    tipstricks = TipsTricks.objects.filter(adminID=idx).order_by('-id')
    context = {'tipstricks': tipstricks, 'admin': adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/show_tipstricks.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_multiple_tipstricks(request):
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        tiptrickids = request.POST.getlist('items[]')

        for tiptrickid in tiptrickids:
            TipsTricks.objects.filter(id=tiptrickid).delete()

        adminUser = AdminUser.objects.get(adminID=idx)
        tipstricks = TipsTricks.objects.filter(adminID=idx).order_by('-id')
        context = {'tipstricks': tipstricks, 'admin': adminUser}
        return render(request, 'vacay/show_tipstricks.html', context)
        # return HttpResponse(serviceids)

    elif request.method == 'GET':
        pass

def edit_tiptrick(request, tiptrick_id):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    tiptrick=TipsTricks.objects.get(id=tiptrick_id)
    context = {'tiptrick': tiptrick}
    # cache.set('search_env', 3)
    return render(request, 'vacay/edit_tipstricks.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edit_tiptrick_process(request, tiptrick_id):
    if request.method == 'POST':

        title = request.POST.get('title', None)
        audience = request.POST.get('audience', None)
        subject = request.POST.get('subject', None)
        callofaction = request.POST.get('callofaction', None)
        owneremail = request.POST.get('owneremail', None)
        description = request.POST.get('description', None)
        survey = request.POST.get('survey', None)
        youtubeurl = request.POST.get('youtubeurl', None)
        admin_free = request.POST.get('afree', None)
        buy_paid = request.POST.get('bpaid', None)
        free = request.POST.get('fr', None)

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        date=strftime("%Y-%m-%d", gmtime())

        tipstricks = TipsTricks.objects.get(id=tiptrick_id)
        tipstricks.adminID = idx
        tipstricks.title = title
        tipstricks.audience = audience
        tipstricks.subject = subject
        tipstricks.description = description
        tipstricks.callofaction = callofaction
        tipstricks.owneremail = owneremail
        tipstricks.postdate = "Updated at: "+date
        tipstricks.survey = survey
        tipstricks.op_admin = admin_free
        tipstricks.op_buy = buy_paid
        tipstricks.op_free = free

        try:
            photo = request.FILES['photo']
            fs = FileSystemStorage()
            filename = fs.save(photo.name, photo)
            uploaded_file_url = fs.url(filename)
            tipstricks.image = settings.URL + uploaded_file_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            videofile = request.FILES['video']
            fs = FileSystemStorage()
            filename = fs.save(videofile.name, videofile)
            video_url = fs.url(filename)
            tipstricks.video = settings.URL + video_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        if youtubeurl is not None:
            tipstricks.youtube = youtubeurl

        tipstricks.save()

        adminUser = AdminUser.objects.get(adminID=idx)
        tipstricks = TipsTricks.objects.filter(adminID=idx).order_by('-id')
        context = {'tipstricks': tipstricks, 'admin':adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_tipstricks.html', context)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def edittiptrickpictures(request, tiptrick_id ):

    global pro_num
    if request.method == 'POST':

        tipstricks = TipsTricks.objects.get(id=tiptrick_id)

        try:
            image = request.FILES['photoa']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            tipstricks.imageA = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photob']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            tipstricks.imageB = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photoc']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            tipstricks.imageC = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photod']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            tipstricks.imageD = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photoe']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            tipstricks.imageE = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        try:
            image = request.FILES['photof']
            fs = FileSystemStorage()
            filename = fs.save(image.name, image)
            image_url = fs.url(filename)
            tipstricks.imageF = settings.URL + image_url
        except MultiValueDictKeyError or FileNotFoundError:
            print("File Not Exist")

        desca = request.POST.get('desca', '')
        descb = request.POST.get('descb', '')
        descc = request.POST.get('descc', '')
        descd = request.POST.get('descd', '')
        desce = request.POST.get('desce', '')
        descf = request.POST.get('descf', '')

        tipstricks.descA=desca
        tipstricks.descB=descb
        tipstricks.descC=descc
        tipstricks.descD=descd
        tipstricks.descE=desce
        tipstricks.descF=descf

        tipstricks.save()

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=idx)
        tipstricks = TipsTricks.objects.filter(adminID=idx).order_by('-id')
        context = {'tipstricks': tipstricks, 'admin':adminUser}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_tipstricks.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_multiple_provider(request):
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        proids = request.POST.getlist('items[]')

        for proid in proids:
            Provider.objects.filter(proid=proid).delete()
            Service.objects.filter(proid=proid).delete()
            Product.objects.filter(proid=proid).delete()
            ProviderSchedule.objects.filter(proid=proid).delete()

        adminUser = AdminUser.objects.get(adminID=idx)
        all_providers = Provider.objects.filter(adminID=idx).order_by('-id')

        context = {'all_providers': all_providers, 'admin': adminUser}

        return render(request, 'vacay/index.html', context)
        # return HttpResponse(proids)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_multiple_service(request):
    global pro_num
    pro_num=request.session['pro_id']
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        serviceids = request.POST.getlist('items[]')

        for serviceid in serviceids:
            Service.objects.filter(serviceid=serviceid).delete()

        provider = Provider.objects.get(proid=pro_num)
        services = Service.objects.filter(proid=pro_num).order_by('-id')
        context = {'services': services, 'provider': provider}
        # cache.set('search_env', 3)
        return render(request, 'vacay/show_services.html', context)
        # return HttpResponse(serviceids)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_multiple_product(request):
    global pro_num
    pro_num = request.session['pro_id']
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        productids = request.POST.getlist('items[]')

        for productid in productids:
            Product.objects.filter(itemid=productid).delete()

        provider = Provider.objects.get(proid=pro_num)
        product = Product.objects.filter(proid=pro_num).order_by('-id')
        context = {'product': product, 'provider': provider}

        return render(request, 'vacay/show_products.html', context)
        # return HttpResponse(serviceids)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_multiple_retail(request):
    global ret
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        bproductids = request.POST.getlist('items[]')

        for bproductid in bproductids:
            BroadmoorProduct.objects.filter(bm_proid=bproductid).delete()
            BroadmoorProductDetail.objects.filter(bm_proid=bproductid).delete()

        ret = 0
        request.session['ret']=0

        adminUser = AdminUser.objects.get(adminID=idx)
        bproducts = BroadmoorProduct.objects.filter(adminID=idx).order_by('-id')
        context = {'bproducts': bproducts, 'admin': adminUser}
        return render(request, 'vacay/show_broadmoor_products.html', context)
        # return HttpResponse(serviceids)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_multiple_employee(request):
    global com
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        employeeids = request.POST.getlist('items[]')

        for employeeid in employeeids:
            Employee.objects.filter(em_id=employeeid).delete()

        com=0
        request.session['com']=0

        adminUser = AdminUser.objects.get(adminID=idx)
        employees = Employee.objects.filter(adminID=idx).order_by('-id')
        context = {'employees': employees, 'admin': adminUser}

        return render(request, 'vacay/show_employees.html', context)
        # return HttpResponse(serviceids)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_multiple_job(request):
    global com
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        jobids = request.POST.getlist('items[]')

        for jobid in jobids:
            Job.objects.filter(job_id=jobid).delete()

        com=0
        request.session['com']=0

        adminUser = AdminUser.objects.get(adminID=idx)
        jobs = Job.objects.filter(adminID=idx).order_by('-id')
        context = {'jobs': jobs, 'admin': adminUser}

        return render(request, 'vacay/show_jobs.html', context)
        # return HttpResponse(serviceids)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delete_multiple_announce(request):
    global com
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        announceids = request.POST.getlist('items[]')

        for announceid in announceids:
            Announce.objects.filter(an_id=announceid).delete()

        com=0
        request.session['com']=0

        adminUser = AdminUser.objects.get(adminID=idx)
        announces = Announce.objects.filter(adminID=idx).order_by('-id')
        context = {'announces': announces, 'admin': adminUser}
        return render(request, 'vacay/show_announcements.html', context)
        # return HttpResponse(serviceids)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def increase_usedbuck(request):

    if request.method == 'POST':

        employeeid = request.POST.get('em_id', None)
        amount = request.POST.get('amount', None)

        employee = Employee.objects.get(em_id=employeeid)
        if employee.em_usedbuck == '':

            if (float(employee.em_givenbuck.replace("$","").replace(",", "")) >= float(str(amount).replace(",", "").replace("$", ""))):
                employee.em_usedbuck = "$" + str(amount).replace(",", "").replace("$", "")
            else:
                resp = {'result_code': '100'}
                return JsonResponse(resp)
        elif employee.em_usedbuck == '0':

            if (float(employee.em_givenbuck.replace("$","").replace(",", "")) >= float(str(amount).replace(",", "").replace("$", ""))):
                employee.em_usedbuck = "$" + str(amount).replace(",", "").replace("$", "")
            else:
                resp = {'result_code': '100'}
                return JsonResponse(resp)
        else:

            if (float(employee.em_givenbuck.replace("$","").replace(",", "")) >= (float(employee.em_usedbuck.replace("$","").replace(",", "")) + float(str(amount).replace(",", "").replace("$", "")))):
                employee.em_usedbuck = "$" + str(float(employee.em_usedbuck.replace("$","").replace(",", "")) + float(str(amount).replace(",", "").replace("$", "")))
            else:
                resp = {'result_code': '100'}
                return JsonResponse(resp)
        employee.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_bucks_data(request, employee_id):

    if request.method == 'GET':

        employee = Employee.objects.filter(em_id=employee_id)
        count = employee.count()
        if count>0:
            serializer = EmployeeSerializer(employee, many=True)
            resp = {'result_code': '0', 'bucks_data': serializer.data, }
            return JsonResponse(resp, status=status.HTTP_200_OK)
        else:
            resp = {'result_code': '113'}
            return JsonResponse(resp)
    elif request.method == 'POST':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_givenbuck(request):

    if request.method == 'POST':

        employeeid = request.POST.get('em_id', None)
        amount = request.POST.get('amount', None)

        employee = Employee.objects.get(em_id=employeeid)

        if (2000.0 >= float(str(amount).replace(",", "").replace("$", ""))) and (float(str(amount).replace(",", "").replace("$", "")) >= float(employee.em_givenbuck.replace("$","").replace(",", ""))):
            employee.em_givenbuck = "$" + str(amount).replace(",", "").replace("$", "")
        elif (2000.0 >= float(str(amount).replace(",", "").replace("$", ""))) and (float(str(amount).replace(",", "").replace("$", "")) < float(employee.em_givenbuck.replace("$","").replace(",", ""))):
            resp = {'result_code': '99'}
            return JsonResponse(resp)
        else:
            resp = {'result_code': '100'}
            return JsonResponse(resp)
        employee.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_account(request):
    if request.method == 'POST':

        stripe_id = request.POST.get('stripe_id', None)
        email = request.POST.get('email', None)
        country = request.POST.get('country', None)

        created_on = strftime("%Y-%m-%d %H:%M:%S", gmtime())

        account = Account()
        account.stripe_id = stripe_id
        account.country = country
        account.email = email
        account.created_on = created_on
        account.status = "Pending verification"

        account.save()

        resp = {
            'result': 'success',
            'account_data':{
                'accountid': account.stripe_id,
                'status': account.status
                }
        }

        # resp = {'result': 'success', 'accountid': stripe_id}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass
@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_account_detail(request):

    if request.method == 'POST':

        email = request.POST.get('email', None)

        accounts = Account.objects.filter(email=email)
        count=accounts.count()
        if count>0:
            resp = {
                'status': 'success',
                'account_data': {
                    'accountid': accounts[0].stripe_id,
                    'status': accounts[0].status
                }
            }
            return JsonResponse(resp, status=status.HTTP_200_OK)
        else:
            resp = {
                'status': 'error',
            }
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_account(request):

    if request.method == 'POST':

        email = request.POST.get('email', None)

        modified_on = strftime("%Y-%m-%d %H:%M:%S", gmtime())

        account = Account.objects.get(email=email)

        account.modified_on = modified_on
        account.status = "Approved"

        account.save()

        resp = {
            'status': 'success'
        }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendPaymentEmail(request):
    if request.method == 'POST':

        fromEmail = request.POST.get('senderEmail', None)
        toEmail = request.POST.get('receiverEmail', None)
        amount = request.POST.get('paidMoney', None)

        fromaddress = 'cayley@vacaycarpediem.com'  # cayley@vacaycarpediem.com
        toaddress = toEmail
        subject = 'VaCay Payment System.'
        body = ''

        html = """\
                <html>
                  <head></head>
                  <body>
                  <img src="https://www.vacayadmin.com/static/vacay/images/vacaylogo.jpg" style="width:80px;height:80px;border-radius: 8%; margin-left:25px;"/>
                    <h3 style="margin-left:10px; color:#02839a;">VaCay payment information</h3>
                  </body>
                </html>
                """

        text = "You received "+amount+" USD from "+fromEmail+"\nWarmly,\nThe VaCay Team\n"

        username = settings.EMAIL_HOST_USER
        password = settings.EMAIL_HOST_PASSWORD
        msg = MIMEMultipart()
        msg['From'] = fromaddress
        msg['To'] = toaddress
        msg['Subject'] = subject
        # msg['Body'] = body
        # message = """From: %s\nTo: %s\nSubject: %s\n\n%s
        # """ % (fromaddress, ", ".join(toaddress), subject, body)

        # body = MIMEText(html, 'html')
        body1 = MIMEText(html, 'html')
        body = MIMEText(text, 'plain')

        msg.attach(body1)
        msg.attach(body)

        try:
            server = smtplib.SMTP('mail.smtp2go.com', 2525)
            server.ehlo()
            server.starttls()
            server.login(username, password)
            server.sendmail(fromaddress, toaddress, msg.as_string())

            server.quit()

            resp = {
                'result': '0'
            }
            return JsonResponse(resp, status=status.HTTP_200_OK)

            # return HttpResponse(toaddress + ': Sent!')
        except:

            resp = {
                'result': '1'
            }
            return JsonResponse(resp)

            # return HttpResponse("failed to send mail")

    elif request.method == 'GET':
        pass

def get_all_service_breakdown(request):

    global pro_num
    pro_num=request.session['pro_id']

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    provider = Provider.objects.get(proid=pro_num)
    services = Service.objects.filter(proid=pro_num)

    processing_fee = 0.2
    servicePercent = float(str(provider.proServicePercent).replace("%","").replace(",", ""))
    managerPercent = 100 - servicePercent

    for service in services:
        totalPrice = float(str(service.proServicePrice).replace("$","").replace(",", "")) - (float(str(service.proServicePrice).replace("$","").replace(",", "")) * processing_fee)
        providerTakeHome = servicePercent * totalPrice * 0.01
        managerTakeHome = totalPrice - providerTakeHome

        service.providerTakeHome = str(providerTakeHome)[0:8]
        service.managerTakeHome = str(managerTakeHome)[0:8]

        service.save()

    processing_fee = str(processing_fee * 100)

    context = {'services': services, 'provider': provider, 'processing_fee':processing_fee, 'servicePercent':servicePercent, 'managerPercent':managerPercent}
    # cache.set('search_env', 3)
    return render(request, 'vacay/service_breakdown.html', context)

def get_all_product_breakdown(request):

    global pro_num
    pro_num=request.session['pro_id']

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    provider = Provider.objects.get(proid=pro_num)
    products = Product.objects.filter(proid=pro_num).order_by('-id')

    processing_fee = 0.2
    providerPercent = float(str(provider.proProductSalePercent).replace("%","").replace(",", ""))
    managerPercent = 100 - providerPercent

    for product in products:
        totalPrice = float(str(product.itemPrice).replace("$","").replace(",","")) - (float(str(product.itemPrice).replace("$","").replace(",","")) * processing_fee)
        providerTakeHome = providerPercent * totalPrice * 0.01
        managerTakeHome = totalPrice - providerTakeHome

        product.providerTakeHome = str(providerTakeHome)[0:8]
        product.managerTakeHome = str(managerTakeHome)[0:8]

        product.save()

    processing_fee = str(processing_fee * 100)

    context = {'products': products, 'provider': provider, 'processing_fee':processing_fee, 'providerPercent':providerPercent, 'managerPercent':managerPercent}
    # cache.set('search_env', 3)
    return render(request, 'vacay/product_breakdown.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_admin_accountid(request):

    if request.method == 'POST':

        email = request.POST.get('proEmail', None)

        provider=Provider.objects.filter(proEmail=email)
        adminid=provider[0].adminID
        adminEmail=AdminUser.objects.get(adminID=adminid).adminEmail
        accounts=Account.objects.filter(email=adminEmail, status='Approved')
        count=accounts.count()
        if count>0:
            accountid=accounts[0].stripe_id
            resp = {
                'result_code': '0',
                'accountid':accountid
            }
            return JsonResponse(resp, status=status.HTTP_200_OK)
        else:
            resp = {
                'result_code': '100'
            }
            return JsonResponse(resp)
        # return HttpResponse(adminEmail)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def register_common_user(request):

    if request.method == 'POST':

        email = request.POST.get('email', None)
        first_name = request.POST.get('first_name', None)
        last_name = request.POST.get('last_name', None)
        age = request.POST.get('age', None)
        phone = request.POST.get('phone_number', None)
        address = request.POST.get('address', None)
        job = request.POST.get('job', None)
        education = request.POST.get('education', None)
        interests = request.POST.get('interests', None)
        relationship = request.POST.get('relationship', None)
        place_name = request.POST.get('place_name', None)
        user_lat = request.POST.get('user_lat', None)
        user_lon = request.POST.get('user_lon', None)
        photo_url = request.POST.get('photo_url', None)
        survey = request.POST.get('survey', None)
        em_millennial = request.POST.get('em_millennial', None)

        employees = Employee.objects.filter(em_email=email)
        if employees.count() > 0:
            em = employees[0]
            em.em_image = photo_url
            em.save()

        users = CommonUser.objects.filter(email=email)
        count = users.count()

        if count ==0:
            commonUser = CommonUser()
            commonUser.email = email
            commonUser.first_name = first_name
            commonUser.last_name = last_name
            commonUser.age = age
            commonUser.phone_number = phone
            commonUser.address = address
            commonUser.job = job
            commonUser.education = education
            commonUser.interests = interests
            commonUser.relationship = relationship
            commonUser.place_name = place_name
            commonUser.user_lat = user_lat
            commonUser.user_lon = user_lon
            commonUser.photo_url = photo_url
            commonUser.survey = survey
            commonUser.em_millennial = em_millennial

            commonUser.save()

            commonUser.userid = commonUser.pk
            commonUser.save()

            user = CommonUser.objects.get(email=email)

            serializer=CommonUserSerializer(user, many=True)

            resp = {'result_code': '0'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

        else:

            commonUser = users[0]
            commonUser.email = email
            commonUser.first_name = first_name
            commonUser.last_name = last_name
            commonUser.age = age
            commonUser.phone_number = phone
            commonUser.address = address
            commonUser.job = job
            commonUser.education = education
            commonUser.interests = interests
            commonUser.relationship = relationship
            commonUser.place_name = place_name
            commonUser.user_lat = user_lat
            commonUser.user_lon = user_lon
            commonUser.photo_url = photo_url
            commonUser.survey = survey
            commonUser.em_millennial = em_millennial

            commonUser.save()

            commonUser.userid = commonUser.pk
            commonUser.save()

            watercoolers = Watercooler.objects.filter(email=email)
            if watercoolers.count() > 0:
                wc = watercoolers[0]
                wc.photoUrl = photo_url
                wc.save()

            comments = Comment.objects.filter(email=email)
            if comments.count() > 0:
                c = comments[0]
                c.photoUrl = photo_url
                c.save()

            employees = Employee.objects.filter(em_email=email)
            if employees.count() > 0:
                em = employees[0]
                em.em_image = photo_url
                em.save()

            user = CommonUser.objects.get(email=email)

            serializer = CommonUserSerializer(user, many=True)

            resp = {'result_code': '101'}
            return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_user_profile(request):

    if request.method == 'POST':

        email = request.POST.get('email', None)

        users = CommonUser.objects.filter(email=email)
        count = users.count()
        if count>0:
            serializer = CommonUserSerializer(users, many=True)
            resp = {'result_code': '0', 'user_profile': serializer.data}
            return JsonResponse(resp, status=status.HTTP_200_OK)
        else:
            resp = {'result_code': '108'}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_all_users(request):

    if request.method == 'POST':

        users = CommonUser.objects.all()

        serializer = CommonUserSerializer(users, many=True)
        resp = {'result_code': '0', 'user_infos': serializer.data, }
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_mail_message(request):

    if request.method == 'POST':

        email = request.POST.get('email', None)

        messages = MailBox.objects.filter(to_mail=email)
        count = messages.count()
        if count>0:
            serializer = MailBoxSerializer(messages, many=True)
            resp = {'result_code': '0', 'mail_infos': serializer.data}
            return JsonResponse(resp, status=status.HTTP_200_OK)
        else:
            resp = {'result_code': '109'}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_sent_message(request):

    if request.method == 'POST':

        email = request.POST.get('email', None)

        messages = MailBox.objects.filter(from_mail=email)
        count = messages.count()
        if count>0:
            serializer = MailBoxSerializer(messages, many=True)
            resp = {'result_code': '0', 'sent_mail_infos': serializer.data}
            return JsonResponse(resp, status=status.HTTP_200_OK)
        else:
            resp = {'result_code': '110'}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def send_message(request):

    if request.method == 'POST':

        from_mail = request.POST.get('from_mail', None)
        to_mail = request.POST.get('to_mail', None)
        text_message = request.POST.get('text_message', None)
        lon_message = request.POST.get('lon_message', None)
        lat_message = request.POST.get('lat_message', None)
        name = request.POST.get('name', None)
        photo_url = request.POST.get('photo_url', None)
        request_date = request.POST.get('request_date', None)
        service = request.POST.get('service', None)
        req_date = request.POST.get('service_reqdate', None)

        mail = MailBox()
        mail.from_mail = from_mail
        mail.to_mail = to_mail
        mail.text_message = text_message
        mail.lon_message = lon_message
        mail.lat_message = lat_message
        mail.name = name
        mail.photo_url = photo_url
        mail.request_date = request_date
        mail.service = service
        mail.service_reqdate = req_date

        mail.save()

        mail.mail_id = mail.pk
        mail.save()

        resp = {'result_code': '0', 'mail_id': mail.pk}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def upload_mail_image(request):

    if request.method == 'POST':

        image = request.FILES['file']

        fs = FileSystemStorage()
        filename = fs.save(image.name, image)
        uploaded_file_url = fs.url(filename)

        mail_id = request.POST.get('mail_id')

        mail = MailBox.objects.get(id=mail_id)
        mail.image_message_url = settings.URL + uploaded_file_url
        mail.save()

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp), status=status.HTTP_200_OK)


    elif request.method == 'GET':

        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendEmailMessage(request):

    if request.method == 'POST':

        mail_id = request.POST.get('mail_id', None)

        mail=MailBox.objects.get(mail_id=mail_id)

        fromaddress = mail.from_mail
        toaddress = mail.to_mail
        subject = 'Hello, how are you?'
        body = ''

        html = """\
                <html>
                  <head></head>
                  <body>
                  <img src="https://www.vacayadmin.com/static/vacay/images/vacaylogo.jpg" style="width:80px;height:80px;border-radius: 8%; margin-left:25px;"/>
                    <h3 style="margin-left:10px; color:#02839a;">VaCay mail information</h3>
                  </body>
                </html>
                """

        text = mail.text_message

        username = settings.EMAIL_HOST_USER
        password = settings.EMAIL_HOST_PASSWORD
        msg = MIMEMultipart()
        msg['From'] = fromaddress
        msg['To'] = toaddress
        msg['Subject'] = subject
        # msg['Body'] = body
        # message = """From: %s\nTo: %s\nSubject: %s\n\n%s
        # """ % (fromaddress, ", ".join(toaddress), subject, body)

        # body = MIMEText(html, 'html')
        body1 = MIMEText(html, 'html')
        body = MIMEText(text, 'plain')

        msg.attach(body1)
        msg.attach(body)

        try:
            server = smtplib.SMTP('mail.smtp2go.com', 2525)
            server.ehlo()
            server.starttls()
            server.login(username, password)
            server.sendmail(fromaddress, toaddress, msg.as_string())

            server.quit()

            resp = {
                'result': '0'
            }
            return JsonResponse(resp, status=status.HTTP_200_OK)

            # return HttpResponse(toaddress + ': Sent!')
        except:

            resp = {
                'result': '1'
            }
            return JsonResponse(resp)

            # return HttpResponse("failed to send mail")

    elif request.method == 'GET':
        pass

def delete_sentMail(request):
    if request.method == 'POST':

        mail_id = request.POST.get('mail_id', None)

        MailBox.objects.get(id=mail_id).delete()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def update_request_message(request):

    if request.method == 'POST':

        mail_id = request.POST.get('mail_id', None)
        sts = request.POST.get('status', None)

        mail = MailBox.objects.get(id=mail_id)

        mail.status = sts

        mail.save()

        resp = {'result_code': '0'}
        return HttpResponse(json.dumps(resp))

    elif request.method == 'GET':
        pass


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendEmailMes(request):

    if request.method == 'POST':

        mail_id = request.POST.get('mail_id', None)
        message = request.POST.get('message', None)

        mail=MailBox.objects.get(mail_id=mail_id)

        fromaddress = mail.from_mail
        toaddress = mail.to_mail
        subject = 'Hi, how do you do?'
        body = ''

        indx = message.find('Thanks')
        mes1 = message[0: indx-1].replace('\n', '<br>')
        mes2 = message[indx: len(message)].replace('\n', '<br>')

        html = """\
                <html>
                  <head></head>
                  <body>
                  <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/vacaylogo.jpg" style="width:90px;height:90px;border-radius: 8%; margin-left:25px;"/></a>
                    <h2 style="margin-left:10px; color:#02839a;">VaCay mail information</h2>
                    <div style="font-size:16px; word-break: break-all; word-wrap: break-word;">
                        {mes1}<br><br>
                        <a href="https://www.vacayalldays.com/proaccept/{mailid}"><label style="color:red; font-size:18px;">Accept</label></a>
                        <a href="https://www.vacayalldays.com/prodecline/{mailid}"><label style="color:red; font-size:18px; margin-left:30px;">Decline</label></a><br><br>
                        {mes2}
                    </div>
                  </body>
                </html>
                """

        html = html.format(mes1=mes1, mes2=mes2, mailid=mail.mail_id)

        username = settings.EMAIL_HOST_USER
        password = settings.EMAIL_HOST_PASSWORD
        msg = MIMEMultipart()
        msg['From'] = fromaddress
        msg['To'] = toaddress
        msg['Subject'] = subject
        # msg['Body'] = body
        # message = """From: %s\nTo: %s\nSubject: %s\n\n%s
        # """ % (fromaddress, ", ".join(toaddress), subject, body)

        # body = MIMEText(html, 'html')
        body1 = MIMEText(html, 'html')
        body = MIMEText(message, 'plain')

        msg.attach(body1)
#        msg.attach(body)

        try:
            server = smtplib.SMTP('mail.smtp2go.com', 2525)
            server.ehlo()
            server.starttls()
            server.login(username, password)
            server.sendmail(fromaddress, toaddress, msg.as_string())

            server.quit()

            resp = {
                'result': '0'
            }
            return JsonResponse(resp, status=status.HTTP_200_OK)

            # return HttpResponse(toaddress + ': Sent!')
        except:

            resp = {
                'result': '1'
            }
            return JsonResponse(resp)

            # return HttpResponse("failed to send mail")

    elif request.method == 'GET':
        pass

def service_multiple(request, service_id):
    service = Service.objects.get(id=service_id)

    context={'object':service}
    return render(request, 'vacay/show_multiple.html', context)

def product_multiple(request, product_id):
    product = Product.objects.get(id=product_id)

    context={'object':product}
    return render(request, 'vacay/show_multiple.html', context)

def retail_multiple(request, broadmoorproduct_id):
    product = BroadmoorProduct.objects.get(id=broadmoorproduct_id)

    context={'object':product}

    return render(request, 'vacay/show_multiple.html', context)

def announce_multiple(request, announce_id):
    announce = Announce.objects.get(id=announce_id)

    context={'object':announce}

    return render(request, 'vacay/show_multiple.html', context)

def provider_picture(request, provider_id):
    provider=Provider.objects.get(id=provider_id)
    context={'provider':provider}
    return render(request, 'vacay/picture_view.html', context)

def announce_picture(request, announce_id):
    announce=Announce.objects.get(id=announce_id)
    context={'announce':announce}
    return render(request, 'vacay/picture_view.html', context)

def service_picture(request, service_id):
    service=Service.objects.get(id=service_id)
    context={'service':service}
    return render(request, 'vacay/picture_view.html', context)

def product_picture(request, product_id):
    product=Product.objects.get(id=product_id)
    context={'product':product}
    return render(request, 'vacay/picture_view.html', context)

def retail_picture(request, broadmoorproduct_id):
    bproduct=BroadmoorProduct.objects.get(id=broadmoorproduct_id)
    context={'bproduct':bproduct}
    return render(request, 'vacay/picture_view.html', context)

def employee_picture(request, employee_id):
    employee=Employee.objects.get(id=employee_id)
    context={'employee':employee}
    return render(request, 'vacay/picture_view.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_media(request):
    if request.method == 'POST':
        item_id = request.POST.get('item_id', None)
        item = request.POST.get('item', None)

        obj = None
        if item == 'service':
            obj = Service.objects.get(id=item_id)
        elif item == 'product':
            obj = Product.objects.get(id=item_id)
        elif item == 'bproduct':
            obj = BroadmoorProduct.objects.get(id=item_id)
        elif item == 'announce':
            obj = Announce.objects.get(id=item_id)

        if obj is not None:
            resp = {
                'result_code': '0',
                'media': {
                    'video_url': obj.video_url,
                    'youtube_url': obj.youtube_url,
                    'image_a': obj.imageA,
                    'image_b': obj.imageB,
                    'image_c': obj.imageC,
                    'image_d': obj.imageD,
                    'image_e': obj.imageE,
                    'image_f': obj.imageF,
                }
            }
            return JsonResponse(resp, status=status.HTTP_200_OK)
        else:
            resp = {'result_code':'1'}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_job_media(request):
    if request.method == 'POST':
        item_id = request.POST.get('item_id', None)
        item = request.POST.get('item', None)

        obj = None
        if item == 'job':
            obj = Job.objects.get(id=item_id)

        if obj is not None:
            resp = {
                'result_code': '0',
                'media': {
                    'video_url': obj.video_url,
                    'youtube_url': obj.youtube_url,
                }
            }
            return JsonResponse(resp, status=status.HTTP_200_OK)
        else:
            resp = {'result_code':'1'}
            return JsonResponse(resp)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_watercooler(request):

    if request.method == 'POST':

        name = request.POST.get('name', None)
        email = request.POST.get('email', None)
        photo = request.POST.get('photo', None)
        company = request.POST.get('company', None)
        category = request.POST.get('category', None)
        content = request.POST.get('content', None)
        link = request.POST.get('link', None)

        watercooler = Watercooler()
        watercooler.name = name
        watercooler.email = email
        watercooler.photoUrl = photo
        watercooler.company = company
        watercooler.category = category
        watercooler.content = content
        watercooler.link = link

        watercooler.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updatewatercooler(request):

    if request.method == 'POST':

        wc_id = request.POST.get('wc_id', '1')
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        photo = request.POST.get('photo', '')
        company = request.POST.get('company', '')
        category = request.POST.get('category', '')
        content = request.POST.get('content', '')
        link = request.POST.get('link', '')

        watercooler = Watercooler.objects.get(id=wc_id)
        watercooler.name = name
        watercooler.email = email
        watercooler.photoUrl = photo
        watercooler.company = company
        watercooler.category = category
        watercooler.content = content
        watercooler.link = link

        watercooler.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_watercooler(request):

    if request.method == 'POST':

        company = request.POST.get('company', None)
        category = request.POST.get('category', None)

        watercooler = Watercooler.objects.filter(company=company, category=category)
        serializer = WatercoolerSerializer(watercooler, many=True)
        resp = {'result_code': '0', 'data': serializer.data}

        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delwatercooler(request):

    if request.method == 'POST':

        wc_id = request.POST.get('wc_id', '1')

        watercooler = Watercooler.objects.get(id=wc_id)
        watercooler.delete()
        resp = {'result_code': '0'}

        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delcomment(request):
    if request.method == 'POST':
        comment_id = request.POST.get('comment_id', '1')

        comment = Comment.objects.get(id=comment_id)
        comment.delete()
        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def add_comment(request):

    if request.method == 'POST':

        info_id = request.POST.get('info_id', '1')
        photo = request.POST.get('photo', '')
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        text = request.POST.get('text', '')
        image = request.POST.get('image', '')

        comments = Comment.objects.filter(email=email, info_id=info_id)
        comment = None
        if comments.count() > 0:
            comment = comments[0]
        else:
            comment = Comment()
        comment.info_id = info_id
        comment.photoUrl = photo
        comment.name = name
        comment.email = email
        comment.text = text
        comment.imageUrl = image

        comment.save()

        resp = {'result_code': '0'}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_comment(request):

    if request.method == 'POST':

        info_id = request.POST.get('info_id', None)

        comments = Comment.objects.filter(info_id=info_id)
        serializer = CommentSerializer(comments, many=True)
        resp = {'result_code': '0', 'data': serializer.data}

        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

def tiptrick_picture(request, tiptrick_id):
    tiptrick=TipsTricks.objects.get(id=tiptrick_id)
    context={'tiptrick':tiptrick}
    return render(request, 'vacay/picture_view.html', context)

def tipstricks_multiple(request, tiptrick_id):
    tipstricks = TipsTricks.objects.get(id=tiptrick_id)

    context={'object':tipstricks}

    return render(request, 'vacay/show_multiple.html', context)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def get_all_jobs_for_sharing(request):

    if request.method == 'POST':

        pickup_dict = {}
        pickup_records = []

        jobs = Job.objects.all()

        for job in jobs:

            jobid = job.job_id
            adminid = job.adminID
            admincompany = AdminUser.objects.get(adminID=adminid).adminCompany
            adminlogo = AdminUser.objects.get(adminID=adminid).adminLogoImageUrl
            name = job.job_name
            req = job.job_req
            department = job.job_department
            location = job.job_location
            description = job.job_description
            postdate = job.job_postdate
            empty = job.job_empty
            survey = job.job_survey

            record = {
                "job_id": jobid,
                "adminID": adminid,
                "adminCompany": admincompany,
                "adminLogoImageUrl": adminlogo,
                "job_name": name,
                "job_req": req,
                "job_department": department,
                "job_location": location,
                "job_description": description,
                "job_postdate": postdate,
                "job_empty": empty,
                "job_survey":survey,
            }

            pickup_records.append(record)

        pickup_dict["job_info"] = pickup_records

        resp = {'result_code': '0', "job_info": pickup_records}
        return JsonResponse(resp, status=status.HTTP_200_OK)

    elif request.method == 'GET':
        pass

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def sendEmEmailfromApp(request):
    if request.method == 'POST':
        em_id = request.POST.get('em_id', None)
        employee = Employee.objects.get(em_id=em_id)
        fromaddress = 'cayley@vacaycarpediem.com'  # cayley@vacaycarpediem.com
        toaddress = employee.em_email
        subject = 'Welcome to VaCay!'
        body = ''

        if "$" in employee.em_givenbuck:
            emgivenbuck = employee.em_givenbuck
        else:
            emgivenbuck = "$" + employee.em_givenbuck

        html = """\
                <html>
                  <head></head>
                  <body>
                  <center>
                    <p style="font-weight:600;font-size:25px; color:#000000; font-family:verdana;">VaCay Welcomes You!</p>
                    <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/vacaylogo.jpg" style="width:100px;height:100px;border-radius: 20%;"/></a>
                    </center>
                    <h2 style="margin-left:20px; color: #e69900;">Your company has signed you up!</h2>
                    <div style="font-size:16px; font-family:verdana; color:black;">
                        VaCay's mission is to help you feel refreshed, rejuvenated and know how appreciated you are after each use!<br><br>
                        Please get to know your <label style="font-weight:600;">colleagues, make new friends</label>, check out the recommended <label style="font-weight:600;">top restaurants</label>, enjoy your on-demand
                        <label style="font-weight:600;">Employee Incentives</label>, have a wonderful time finding your colleagues to <label style="font-weight:600;">play activities and explore with</label>
                         and learn about your company's <label style="font-weight:600;">Jobs</label> and
                        <label style="font-weight:600;">Announcements</label>!<br><br>
                        Your company has given you {givenbuck} <label style="font-weight:600; color:#e69900;">VaCay Bucks</label> <img src="https://www.vacayadmin.com/static/vacay/images/bucklogo.png" style="width:150px; height:70px;"><br><br>
                        which can be used for the <label style="font-weight:600;">On-Demand Employee Appreciation Services</label>. <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/beautyservicebackground.jpg" style="width:90px; height:90px;"></a><br><br>
                        Your username is <a href="mailto:{username}">{username}</a> and we will send you another email with your password and how to login.<br><br>
                        Have a wonderful day!<br><br>
                        Warmly,<br><br>
                        The VaCay Team<br><br>
                        <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/logo.jpg" style="width:140px; height:70px;"></a>
                    </div>
                  </body>
                </html>
                """
        if emgivenbuck.endswith('.0') or emgivenbuck.endswith('.00'):
            emgivenbuck = emgivenbuck.replace('.0', '').replace('.00', '')
        html = html.format(givenbuck=emgivenbuck, username=employee.em_email)

        html2 = """\
                    <html>
                      <head></head>
                      <body>
                      <center>
                        <p style="font-weight:600;font-size:25px; color:#000000; font-family:verdana;">VaCay Welcomes You!</p>
                        <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/vacaylogo.jpg" style="width:100px;height:100px;border-radius: 20%;"/></a>
                        </center>
                        <h2 style="margin-left:20px; color: #e69900;">Your company has signed you up!</h2>
                        <div style="font-size:16px; color:black;">
                            Please click <a href="https://www.vacayalldays.com/employee_login_page?csrfmiddlewaretoken=38zU9eycA2RneFE4UqoYpggV6NLwWq7g2RgluAvKMG6JxvCfuE9hens0TaFbun4U"><label style="color:#e69900; font-weight:600;">here</label></a> to login.<br><br>
                            If you logout, you'll need to click "Employee" <a href="https://www.vacayalldays.com"><button style="width:150px; height:40px; text-align:center; font-size:16px; font-weight:600; color:white; background:green; border-radius:50px;">Employee</button></a> to login again.<br><br>
                            You'll enter in your username <a href="mailto:{username}"><label style="font-weight:500;">{username}</label></a> and the password <label style="font-weight:600;">{password}</label>.<br><br>
                            You'll also be able to review any FAQs here <a href="https://www.vacayalldays.com"><button style="width:150px; height:40px; text-align:center; font-size:16px; font-weight:600; color:white; background:green; border-radius:50px;">FAQs</button></a> and any Jobs, here <a href="https://www.vacayalldays.com"><button style="width:150px; height:40px; text-align:center; font-size:16px; font-weight:600; color:white; background:green; border-radius:50px;">Jobs</button></a>.<br><br>
                            Thank you!<br><br>
                            Warmly,<br>
                            The VaCay Team<br><br>
                            <a href="https://www.vacayalldays.com"><img src="https://www.vacayadmin.com/static/vacay/images/logo.jpg" style="width:140px; height:70px;"></a>
                        </div>
                      </body>
                    </html>
                    """
        html2 = html2.format(password=employee.em_password, username=employee.em_email)

        html1 = """\
            <html>
              <head></head>
              <body>
                <p>Please visit to <a href="http://www.vacaycarpediem.com" style="font-size:17px; font-style:italic;">VaCay</a>
                </p>
              </body>
            </html>
            """

        text = "VaCay's mission is to help you feel refreshed, rejuvenated and know how appreciated you are after each use!\nPlease get to know your colleagues, " \
               "make new friends, check out the recommended top restaurants, enjoy your on-demand Employee Incentives, have a wonderful time finding your colleagues to play activities and " \
               "explore with and learn about your company's Jobs and Announcements!\n" \
               "Your company has given you " + emgivenbuck + " VaCay bucks which can be used for the On-Demand Employee Appreciation Services.\n" \
                                                             "Your username is " + "manish@vacaycarpediem.com" + " and we will send you another email with your password and how to login.\n" \
                                                                                                                 "Have a wonderful day!\nWarmly,\nThe VaCay Team\n"

        text2 = "Here is your password: " + employee.em_password + "\nWarmly,\nThe VaCay Team\nBelow are some instructions to help you best use VaCay. Please click on the link to learn more. \n"

        username = settings.EMAIL_HOST_USER
        password = settings.EMAIL_HOST_PASSWORD
        msg = MIMEMultipart()
        msg['From'] = fromaddress
        msg['To'] = toaddress
        msg['Subject'] = subject
        # msg['Body'] = body
        # message = """From: %s\nTo: %s\nSubject: %s\n\n%s
        # """ % (fromaddress, ", ".join(toaddress), subject, body)

        # body = MIMEText(html, 'html')
        body1 = MIMEText(html, 'html')
        body2 = MIMEText(html1, 'html')
        body4 = MIMEText(html2, 'html')
        body = MIMEText(text, 'plain')
        body3 = MIMEText(text2, 'plain')

        msg.attach(body1)
        # msg.attach(body)
        # msg.attach(body2)

        try:
            server = smtplib.SMTP('mail.smtp2go.com', 2525)
            server.ehlo()
            server.starttls()
            server.login(username, password)
            server.sendmail(fromaddress, toaddress, msg.as_string())

            msg2 = MIMEMultipart()
            msg2['From'] = fromaddress
            msg2['To'] = toaddress
            msg2['Subject'] = subject

            # msg2.attach(body1)
            # msg2.attach(body3)
            msg2.attach(body4)

            server.sendmail(fromaddress, toaddress, msg2.as_string())
            server.quit()

            employee.em_status = '1'
            employee.save()

            resp = {'result_code':'0'}
            return JsonResponse(resp, status=status.HTTP_200_OK)

            # return HttpResponse(toaddress + ': Sent!')
        except:
            resp = {'result_code':'1'}
            return JsonResponse(resp)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def allmailsforserviceprovideracceptordecline(request):
    if request.method == 'POST':
        mails = MailBox.objects.all()
        serializer = MailBoxSerializer(mails, many=True)
        resp = {'result_code': '0', 'maildata': serializer.data}
        return JsonResponse(resp, status=status.HTTP_200_OK)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def loadphotofromcaysees(request):
    if request.method == 'POST':

        idx = request.POST.get('id', 1)
        image = request.FILES['file']

        fs = FileSystemStorage()

        imgs = Img.objects.filter(admin_id=idx)
        if imgs.count() > 0:
            img = imgs[0]
            fs.delete(img.image_url)
            filename = fs.save("myimage.jpg", image)
            image_url = fs.url(filename)
            img.image_url = filename
            img.save()
        else:
            img = Img()
            img.admin_id = idx
            filename = fs.save("myimage.jpg", image)
            image_url = fs.url(filename)
            img.image_url = filename
            img.save()

        resp = {'result_code':'0', 'image': image_url}

        return JsonResponse(resp)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def amazon(request):
    if request.method == 'POST':
        image = request.FILES['photo']
        fs = FileSystemStorage()

        imgs = Img.objects.filter(admin_id=1)
        if imgs.count() > 0:
            img = imgs[0]
            fs.delete(img.image_url)
            filename = fs.save("myimage.jpg", image)
            image_url = fs.url(filename)
            img.image_url = filename
            img.save()
        else:
            img = Img()
            img.admin_id = '1'
            filename = fs.save("myimage.jpg", image)
            image_url = fs.url(filename)
            img.image_url = filename
            img.save()

        img_url = settings.URL + image_url

        amazon_url = 'http://www.amazon.com/s/ref=nb_sb_noss?url=search-alias%3Daps&field-keywords='

        url = 'https://camfind.p.mashape.com/image_requests'
        res = image_recognize(url=url, key=settings.MASHAPE_API_KEY, image_url=img_url)
        sts = res.get('status')
        token = res.get('token')
        # return HttpResponse(json.dumps(res))
        if sts == 'completed':
            keyword = res.get('name')
            keyword = keyword.replace(' ', '+')
            amazon_url = amazon_url + keyword
            return redirect(amazon_url)
        elif sts == 'not completed':
            time.sleep(20)
            url = 'https://camfind.p.mashape.com/image_responses/'
            res2 = image_response(url=url, key=settings.MASHAPE_API_KEY, token=token)
            sts2 = res2.get('status')
            #  return HttpResponse(json.dumps(res2))
            if sts2 == 'completed':
                keyword = res2.get('name')
                keyword = keyword.replace(' ', '+')
                amazon_url = amazon_url + keyword
                return redirect(amazon_url)
            else:
                return HttpResponse("We couldn\'t recognize this image. Please try again...")
        else:
            return HttpResponse("We couldn\'t recognize this image. Please try again...")

    elif request.method == 'GET':
        pass

import requests

def image_recognize(url, key, image_url):
    headers = {
        'X-Mashape-Key': key
    }
    data = {
        'focus[x]': '480',
        'focus[y]': '640',
        'image_request[remote_image_url]': image_url,
        'image_request[language]': 'en',
        'image_request[locale]': 'en_US',
        'image_request[altitude]': '27.912109375',
        'image_request[latitude]': '35.8714220766008',
        'image_request[longitude]': '14.3583203002251'
    }
    response = requests.post(url, headers=headers, data=data)
    return response.json()

def image_response(url, key, token):
    url = url + token
    headers = {
        'X-Mashape-Key': key,
        'Accept': 'application/json'
    }
    response = requests.get(url, headers=headers)
    return response.json()


def search(request):
    return render(request, 'vacay/product_recognition.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def postAdminPhoto(request):
    if request.method == 'POST':
        admin_id = request.POST.get('admin_id', '1')
        photo_url = request.POST.get('photo_url', '')
        adminuser = AdminUser.objects.get(id=admin_id)
        fname = adminuser.adminImageUrl
        adminuser.adminImageUrl = photo_url
        adminuser.save()

        fs = FileSystemStorage()
        fname = fname.replace(settings.URL + '/pictures/', '')
        fs.delete(fname)

        resp = {'result_code':'0'}
        return JsonResponse(resp)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def editUserPhoto(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        photo_url = request.POST.get('photo_url', '')

        users = CommonUser.objects.filter(email=email)
        if users.count() > 0:
            user = users[0]
            user.photo_url = photo_url
            user.save()

        emps = Employee.objects.filter(em_email=email)
        if emps.count() > 0:
            emp = emps[0]
            emp.em_image = photo_url
            emp.save()

        watercoolers = Watercooler.objects.filter(email=email)
        if watercoolers.count() > 0:
            wc = watercoolers[0]
            wc.photoUrl = photo_url
            wc.save()

        comments = Comment.objects.filter(email=email)
        if comments.count() > 0:
            c = comments[0]
            c.photoUrl = photo_url
            c.save()

        resp = {'result_code':'0'}
        return JsonResponse(resp)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def editProviderProfilePhoto(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        photo_url = request.POST.get('photo_url', '')

        provs = Provider.objects.filter(proEmail=email)
        if provs.count() > 0:
            pro = provs[0]
            pro.proProfileImageUrl = photo_url
            pro.save()

        resp = {'result_code':'0'}
        return JsonResponse(resp)

def employeewatercoolers(request):
    global com
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    com = 0
    request.session['com'] = 0

    adminUser = AdminUser.objects.get(adminID=idx)
    wcs = Watercooler.objects.filter(company=adminUser.adminCompany).order_by('-id')
    for wc in wcs:
        comments = Comment.objects.filter(info_id=wc.pk)
        wc.comments = str(comments.count())
    context = {'watercoolers': wcs, 'admin':adminUser}
    return render(request, 'vacay/watercooler_list.html', context)


def wcdetail(request,wc_id):
    global com
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    com = 0
    request.session['com'] = 0

    adminUser = AdminUser.objects.get(adminID=idx)
    watercooler = Watercooler.objects.get(id=wc_id)
    comments = Comment.objects.filter(info_id=wc_id)

    context = {'comments': comments, 'watercooler':watercooler}
    return render(request, 'vacay/watercooler_comments.html', context)

def delwc(request,wc_id):
    global com
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    com = 0
    request.session['com'] = 0

    watercooler = Watercooler.objects.get(id=wc_id)
    watercooler.delete()
    comments = Comment.objects.filter(info_id=wc_id)
    for comment in comments:
        comment.delete()

    adminUser = AdminUser.objects.get(adminID=idx)

    wcs = Watercooler.objects.filter(company=adminUser.adminCompany).order_by('-id')
    context = {'watercoolers': wcs, 'admin':adminUser}
    return render(request, 'vacay/watercooler_list.html', context)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delmultiwcs(request):
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        wcids = request.POST.getlist('items[]')

        for wcid in wcids:
            watercooler = Watercooler.objects.get(id=wcid)
            watercooler.delete()
            comments = Comment.objects.filter(info_id=wcid)
            for comment in comments:
                comment.delete()

        adminUser = AdminUser.objects.get(adminID=idx)

        wcs = Watercooler.objects.filter(company=adminUser.adminCompany).order_by('-id')
        context = {'watercoolers': wcs, 'admin':adminUser}
        return render(request, 'vacay/watercooler_list.html', context)

    elif request.method == 'GET':
        pass


def delemcomment(request,comment_id, wc_id):
    global com
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    com = 0
    request.session['com'] = 0

    comment = Comment.objects.get(id=comment_id)
    comment.delete()
    adminUser = AdminUser.objects.get(adminID=idx)
    watercooler = Watercooler.objects.get(id=wc_id)
    comments = Comment.objects.filter(info_id=wc_id)

    context = {'comments': comments, 'watercooler':watercooler}
    return render(request, 'vacay/watercooler_comments.html', context)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def delmultiemcomments(request):
    if request.method == 'POST':

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        commentids = request.POST.getlist('items[]')

        for cid in commentids:
            comment = Comment.objects.get(id=cid)
            comment.delete()

        adminUser = AdminUser.objects.get(adminID=idx)
        watercooler = Watercooler.objects.get(id=wc_id)
        comments = Comment.objects.filter(info_id=wc_id)

        context = {'comments': comments, 'watercooler':watercooler}
        return render(request, 'vacay/watercooler_comments.html', context)

    elif request.method == 'GET':
        pass


def emprofile(request,wc_id):
    global com
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    com = 0
    request.session['com'] = 0

    watercooler = Watercooler.objects.get(id=wc_id)
    users = CommonUser.objects.filter(email=watercooler.email)
    if users.count() > 0:
        user = users[0]
        user.relationship = '-' + user.relationship.replace('\ncommon', 'common').replace('\n', '\n-')
        user.interests = "-" + user.interests.replace("{", "").replace("}", "").replace("\",","\n-").replace("\"", "").replace("[", "").replace("]", "")
        return render(request, 'vacay/employee_profile.html', {'user':user})
    else:
        return redirect('/home')


def tomessageoptions(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    announces = Announce.objects.filter(adminID=idx)
    return render(request, 'vacay/message_options.html', {'announces':announces})


def emsignedupforan(request):
    announceid = request.GET['an_id']
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    anviews = AnnounceView.objects.filter(an_id=announceid, is_signup='yes')
    employeeList = []
    for anview in anviews:
        employees=Employee.objects.filter(adminID=idx, em_id=anview.em_id)
        employeeList.append(employees[0])
    context = {'employees': employeeList, 'admin':adminUser}
    # cache.set('search_env', 3)
    return render(request, 'vacay/view_anemployees.html', context)


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def messagebygrouping(request):
    if request.method == 'POST':
        cohort = request.POST.get('cohort', '')
        an_id = request.POST.get('announcement', '')
        subject = request.POST.get('subject', '')
        body = request.POST.get('body', '')

        run = request.POST.get('run', '')
        golf = request.POST.get('golf', '')
        tennis = request.POST.get('tennis', '')
        ski = request.POST.get('ski', '')
        biking = request.POST.get('biking', '')
        fishing = request.POST.get('fishing', '')
        surfing = request.POST.get('surfing', '')
        exploring = request.POST.get('exploring', '')

        rwalk = request.POST.get('walk', '')
        r30mins = request.POST.get('30mins', '')
        r1mile = request.POST.get('1mile', '')
        r5miles = request.POST.get('5miles', '')
        r10miles = request.POST.get('10miles', '')

        gbeginner = request.POST.get('beginner', '')
        gintermediate = request.POST.get('intermediate', '')
        gadvanced = request.POST.get('advanced', '')
        gscramble = request.POST.get('scramble', '')

        t3p = request.POST.get('3p', '')
        t4p = request.POST.get('4p', '')
        t5p = request.POST.get('5p', '')
        t6p = request.POST.get('6p', '')
        twtf = request.POST.get('wtf', '')

        shot = request.POST.get('hot', '')
        sbunny = request.POST.get('bunny', '')
        sgreens = request.POST.get('greens', '')
        sblues = request.POST.get('blues', '')
        sblacks = request.POST.get('blacks', '')
        sdouble = request.POST.get('double', '')
        snotafraid = request.POST.get('notafraid', '')

        bshort = request.POST.get('short', '')
        blong = request.POST.get('long', '')
        bmountain = request.POST.get('mountain', '')
        broad = request.POST.get('road', '')

        ffly = request.POST.get('fly', '')
        flake = request.POST.get('lake', '')
        focean = request.POST.get('ocean', '')

        surocean = request.POST.get('oceansurfing', '')
        surlake = request.POST.get('lakesurfing', '')
        kitesurfing = request.POST.get('kitesurfing', '')

        emuseums = request.POST.get('museums', '')
        ecitytours = request.POST.get('citytours', '')
        enature = request.POST.get('nature', '')
        eart = request.POST.get('art', '')
        econcerts = request.POST.get('concerts', '')
        electures = request.POST.get('lectures', '')

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        employees = Employee.objects.filter(adminID=idx)
        if employees.count() == 0:
            return render(request, 'vacay/result.html', {'response':'You don\'t have any of your employees.'})

        data = []

        if run != '':
            runList = []
            if rwalk != '' and rwalk is not None:
                runList.append(rwalk)
            if r30mins != '' and r30mins is not None:
                runList.append(r30mins)
            if r1mile != '' and r1mile is not None:
                runList.append(r1mile)
            if r5miles != '' and r5miles is not None:
                runList.append(r5miles)
            if r10miles != '' and r10miles is not None:
                runList.append(r10miles)
            if len(runList) > 0:
                runObj = {
                    "Run":str(runList).replace('\'', '')
                }
                data.append(runObj)

        if golf != '':
            golfList = []
            if gbeginner != '' and gbeginner is not None:
                golfList.append(gbeginner)
            if gintermediate != '' and gintermediate is not None:
                golfList.append(gintermediate)
            if gadvanced != '' and gadvanced is not None:
                golfList.append(gadvanced)
            if gscramble != '' and gscramble is not None:
                golfList.append(gscramble)
            if len(golfList) > 0:
                golfObj = {
                    "Golf":str(golfList).replace('\'', '')
                }
                data.append(golfObj)

        if tennis != '':
            tennisList = []
            if t3p != '' and t3p is not None:
                tennisList.append(t3p)
            if t4p != '' and t4p is not None:
                tennisList.append(t4p)
            if t5p != '' and t5p is not None:
                tennisList.append(t5p)
            if t6p != '' and t6p is not None:
                tennisList.append(t6p)
            if twtf != '' and twtf is not None:
                tennisList.append(twtf)
            if len(tennisList) > 0:
                tennisObj = {
                    "Tennis":str(tennisList).replace('\'', '')
                }
                data.append(tennisObj)

        if ski != '':
            skiList = []
            if shot != '' and shot is not None:
                skiList.append(shot)
            if sbunny != '' and sbunny is not None:
                skiList.append(sbunny)
            if sgreens != '' and sgreens is not None:
                skiList.append(sgreens)
            if sblues != '' and sblues is not None:
                skiList.append(sblues)
            if sblacks != '' and sblacks is not None:
                skiList.append(sblacks)
            if sdouble != '' and sdouble is not None:
                skiList.append(sdouble)
            if snotafraid != '' and snotafraid is not None:
                skiList.append(snotafraid)
            if len(skiList) > 0:
                skiObj = {
                    "Ski & Snowboard":str(skiList).replace('\'', '').replace('Lodgeskier', 'Lodge\'skier\'')
                }
                data.append(skiObj)

        if biking != '':
            bikingList = []
            if bshort != '' and bshort is not None:
                bikingList.append(bshort)
            if blong != '' and blong is not None:
                bikingList.append(blong)
            if bmountain != '' and bmountain is not None:
                bikingList.append(bmountain)
            if broad != '' and broad is not None:
                bikingList.append(broad)
            if len(bikingList) > 0:
                bikingObj = {
                    "Biking":str(bikingList).replace('\'', '')
                }
                data.append(bikingObj)

        if fishing != '':
            fishingList = []
            if ffly != '' and ffly is not None:
                fishingList.append(ffly)
            if flake != '' and flake is not None:
                fishingList.append(flake)
            if focean != '' and focean is not None:
                fishingList.append(focean)
            if len(fishingList) > 0:
                fishingObj = {
                    "Fishing":str(fishingList).replace('\'', '')
                }
                data.append(fishingObj)

        if surfing != '':
            surfingList = []
            if surocean != '' and surocean is not None:
                surfingList.append(surocean)
            if surlake != '' and surlake is not None:
                surfingList.append(surlake)
            if kitesurfing != '' and kitesurfing is not None:
                surfingList.append(kitesurfing)
            if len(surfingList) > 0:
                surfingObj = {
                    "Surfing/Kitesurfing":str(surfingList).replace('\'', '')
                }
                data.append(surfingObj)

        if exploring != '':
            exploringList = []
            if emuseums != '' and emuseums is not None:
                exploringList.append(emuseums)
            if ecitytours != '' and ecitytours is not None:
                exploringList.append(ecitytours)
            if enature != '' and enature is not None:
                exploringList.append(enature)
            if eart != '' and eart is not None:
                exploringList.append(eart)
            if econcerts != '' and econcerts is not None:
                exploringList.append(econcerts)
            if electures != '' and electures is not None:
                exploringList.append(electures)
            if len(exploringList) > 0:
                exploringObj = {
                    "Exploring":str(exploringList)
                        .replace('\'', '')
                }
                data.append(exploringObj)

        if len(data) == 0 and cohort == '' and an_id == '':
            # return render(request, 'vacay/result.html', {'response':'Please select at least one option.'})
            toemails = []
            for em in employees:
                toemails.append(em.em_email)
            response = sendMessage(idx, toemails, subject, body)
            return render(request, 'vacay/result.html', {'response':'Sent messages to the employees.'})


        emidList = []

        if cohort != '':
            for em in employees:
                eml = em.em_email
                users = CommonUser.objects.filter(email=eml)
                if users.count() > 0:
                    user = users[0]
                    if user.relationship == cohort:
                        emidList.append(em.em_id)

        if an_id != '':
            anviews = AnnounceView.objects.filter(an_id=an_id, is_signup='yes')
            if anviews.count() > 0:
                for anview in anviews:
                    employees=Employee.objects.filter(adminID=idx, em_id=anview.em_id)
                    em = employees[0]
                    emidList.append(em.em_id)

        if len(data) > 0:

            for em in employees:
                eml = em.em_email
                users = CommonUser.objects.filter(email=eml)
                if users.count() > 0:
                    user = users[0]
                    if user.interests != '':

                        if run != '':
                            if rwalk != '' and rwalk is not None:
                                if rwalk in user.interests:
                                    emidList.append(em.em_id)
                            if r30mins != '' and r30mins is not None:
                                if r30mins in user.interests:
                                    emidList.append(em.em_id)
                            if r1mile != '' and r1mile is not None:
                                if r1mile in user.interests:
                                    emidList.append(em.em_id)
                            if r5miles != '' and r5miles is not None:
                                if r5miles in user.interests:
                                    emidList.append(em.em_id)
                            if r10miles != '' and r10miles is not None:
                                if r10miles in user.interests:
                                    emidList.append(em.em_id)

                        if golf != '':
                            if gbeginner != '' and gbeginner is not None:
                                if gbeginner in user.interests:
                                    emidList.append(em.em_id)
                            if gintermediate != '' and gintermediate is not None:
                                if gintermediate in user.interests:
                                    emidList.append(em.em_id)
                            if gadvanced != '' and gadvanced is not None:
                                if gadvanced in user.interests:
                                    emidList.append(em.em_id)
                            if gscramble != '' and gscramble is not None:
                                if gscramble in user.interests:
                                    emidList.append(em.em_id)

                        if tennis != '':
                            if t3p != '' and t3p is not None:
                                if t3p in user.interests:
                                    emidList.append(em.em_id)
                            if t4p != '' and t4p is not None:
                                if t4p in user.interests:
                                    emidList.append(em.em_id)
                            if t5p != '' and t5p is not None:
                                if t5p in user.interests:
                                    emidList.append(em.em_id)
                            if t6p != '' and t6p is not None:
                                if t6p in user.interests:
                                    emidList.append(em.em_id)
                            if twtf != '' and twtf is not None:
                                if twtf in user.interests:
                                    emidList.append(em.em_id)

                        if ski != '':
                            if shot != '' and shot is not None:
                                if shot in user.interests:
                                    emidList.append(em.em_id)
                            if sbunny != '' and sbunny is not None:
                                if sbunny in user.interests:
                                    emidList.append(em.em_id)
                            if sgreens != '' and sgreens is not None:
                                if sgreens in user.interests:
                                    emidList.append(em.em_id)
                            if sblues != '' and sblues is not None:
                                if sblues in user.interests:
                                    emidList.append(em.em_id)
                            if sblacks != '' and sblacks is not None:
                                if sblacks in user.interests:
                                    emidList.append(em.em_id)
                            if sdouble != '' and sdouble is not None:
                                if sdouble in user.interests:
                                    emidList.append(em.em_id)
                            if snotafraid != '' and snotafraid is not None:
                                if snotafraid in user.interests:
                                    emidList.append(em.em_id)

                        if biking != '':
                            if bshort != '' and bshort is not None:
                                if bshort in user.interests:
                                    emidList.append(em.em_id)
                            if blong != '' and blong is not None:
                                if blong in user.interests:
                                    emidList.append(em.em_id)
                            if bmountain != '' and bmountain is not None:
                                if bmountain in user.interests:
                                    emidList.append(em.em_id)
                            if broad != '' and broad is not None:
                                if broad in user.interests:
                                    emidList.append(em.em_id)

                        if fishing != '':
                            if ffly != '' and ffly is not None:
                                if ffly in user.interests:
                                    emidList.append(em.em_id)
                            if flake != '' and flake is not None:
                                if flake in user.interests:
                                    emidList.append(em.em_id)
                            if focean != '' and focean is not None:
                                if focean in user.interests:
                                    emidList.append(em.em_id)

                        if surfing != '':
                            if surocean != '' and surocean is not None:
                                if surocean in user.interests:
                                    emidList.append(em.em_id)
                            if surlake != '' and surlake is not None:
                                if surlake in user.interests:
                                    emidList.append(em.em_id)
                            if kitesurfing != '' and kitesurfing is not None:
                                if kitesurfing in user.interests:
                                    emidList.append(em.em_id)

                        if exploring != '':
                            if emuseums != '' and emuseums is not None:
                                if emuseums in user.interests:
                                    emidList.append(em.em_id)
                            if ecitytours != '' and ecitytours is not None:
                                if ecitytours in user.interests:
                                    emidList.append(em.em_id)
                            if enature != '' and enature is not None:
                                if enature in user.interests:
                                    emidList.append(em.em_id)
                            if eart != '' and eart is not None:
                                if eart in user.interests:
                                    emidList.append(em.em_id)
                            if econcerts != '' and econcerts is not None:
                                if econcerts in user.interests:
                                    emidList.append(em.em_id)
                            if electures != '' and electures is not None:
                                if electures in user.interests:
                                    emidList.append(em.em_id)

        toemails = []
        toids = []
        if len(emidList) > 0:
            for emid in emidList:
                if emid not in toids:
                    toids.append(emid)
                    em = Employee.objects.get(id=emid)
                    toemails.append(em.em_email)
        else:
            return render(request, 'vacay/result.html', {'response':'No exists employee...'})


        response = sendMessage(idx, toemails, subject, body)

        # return redirect('employees/2/')
        return render(request, 'vacay/result.html', {'response':'Sent messages to the employees.'})


def sendMessage(adminid, toemails, subject, message):
    adminUser = AdminUser.objects.get(adminID=adminid)
    fromemail = adminUser.adminEmail

    msg = EmailMultiAlternatives(subject, message, fromemail, toemails)
    msg.send(fail_silently=False)

    msgText = subject + '\n' + message
    fromemail = fromemail.replace('.com', '').replace(".","ddoott")
    db = firebase.database()
    data = {
        "message": msgText,
        "image": '',
        "video": '',
        "lat": '',
        "lon": '',
        "time":str(int(round(time.time() * 1000))),
        "user": fromemail,
    }

    data2 = {
        "sender": fromemail,
        "senderName": adminUser.adminName,
        "msg": msgText,
        "senderPhoto": adminUser.adminImageUrl,
    }

    data3 = {
        "sender": adminUser.adminName,
        "time": str(int(round(time.time() * 1000))),
        "online": 'online',
    }

    for toemail in toemails:
        toemail = toemail.replace('.com', '').replace(".","ddoott")
        db.child("messages").child(fromemail + '_' + toemail).push(data)
        db.child("messages").child(toemail + '_' + fromemail).push(data)
        db.child("notification").child(toemail + '/' + fromemail).remove()
        db.child("notification").child(toemail + '/' + fromemail).push(data2)
        db.child("status").child(toemail + '/' + fromemail).remove()
        db.child("status").child(toemail + '/' + fromemail).push(data3)

        if toemail == toemails[-1]:
            return True

    return False



def tochat(request):
    em_id = request.GET['em_id']

    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)

    em = Employee.objects.get(em_id=em_id)
    friends = CommonUser.objects.filter(email=em.em_email)
    if friends.count() > 0:
        friend = friends[0]
        return render(request, 'vacay/chat.html', {'me': adminUser, 'friend': friend})
    else:
        return render(request, 'vacay/result.html', {'response':'This employee hasn\'t logged in yet.'})


def get_notifications(request):
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')

    if idx is None:
        return redirect('/logout')

    adminUser = AdminUser.objects.get(adminID=idx)
    return render(request, 'vacay/notification.html', {'me_email': adminUser.adminEmail})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def chat_page(request):
    if request.method == 'POST':
        friend_email = request.POST.get('friend_email', None)
        friend_name = request.POST.get('friend_name', None)
        friend_photo = request.POST.get('friend_photo', None)
        if not '.com' in friend_email:
            friend_email = friend_email.replace('ddoott', '.') + '.com'

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        if idx is None:
            return redirect('/logout')

        adminUser = AdminUser.objects.get(adminID=idx)

        friends = CommonUser.objects.filter(email=friend_email)
        if friends.count() > 0:
            friend = friends[0]
            return render(request, 'vacay/chat.html', {'me': adminUser, 'friend': friend})
        else:
            return render(request, 'vacay/result.html', {'response':'This employee hasn\'t logged in yet.'})




































































